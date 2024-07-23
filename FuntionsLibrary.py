# -----------------------------------------------------------------------------
# ------------------------------ LIBRERIA RIESGO ------------------------------
# -----------------------------------------------------------------------------
"""
-------------------------------------------------------------------------------
Este script contiene las funciones que permiten procesar los resultados de 
riesgo
---------------------------- Autor: Daniela Novoa -----------------------------
"""
#%% ====== Importar librerias =================================================
""" 
-------------------------------------------------------------------------------
*** Ubicar secciones para guardar las librerias
-------------------------------------------------------------------------------
"""
# -------- Librerias Interfaz>> Tkinter ---------------------------------------
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
# -------- Librerias Graficos Interfaz>> Tkinter ------------------------------
from PIL import Image, ImageTk
import matplotlib.ticker as ticker
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
# -------- Librerias Directorios ----------------------------------------------
import os
import glob
import zipfile
import io
# -------- Librerias procesamiento de datos -----------------------------------
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
# -------- Librerias para generar las tablas de resumen -----------------------
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
# -------- Librerias Graficos Interfaz>> Tkinter ------------------------------
import matplotlib.ticker as ticker
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

import geopandas as gpd
import contextily as ctx
from matplotlib.colors import Normalize
from matplotlib.ticker import FuncFormatter
from matplotlib.colors import ListedColormap
from matplotlib import cm
import matplotlib.offsetbox as offsetbox
from matplotlib.patches import Rectangle
import re

from tkinter import filedialog
from tkinter import messagebox as tk_messagebox
from reportlab.pdfgen import canvas
import PyPDF2


#%% ====== Image ==============================================================
def Label_Image(image_name, lw, lh, container,bg_color,rx,ry):
    image = Image.open(os.path.join(os.getcwd(),"icon") + image_name).convert("RGBA")
    image = image.resize((lw, lh), Image.LANCZOS)
    image = ImageTk.PhotoImage(image)
    label = tk.Label(container, image=image, bd=0, bg=bg_color)
    label.image = image
    label.place(relx=rx, rely=ry, anchor=tk.CENTER)
    return label
#%% ====== Rectangle with rounded corners =====================================
def rec_redond(canvas, x1, y1, x2, y2, radio_esquinas, color):
    canvas.create_rectangle(x1 + radio_esquinas, y1, x2 - (radio_esquinas), y2 - radio_esquinas, fill=color, outline=color, width=0)
    canvas.create_rectangle(x1 + radio_esquinas, y1, x2 - (radio_esquinas), y2+1, fill=color, outline=color, width=0)
    canvas.create_rectangle(x1, y1 + radio_esquinas, x2+20, y2 - radio_esquinas, fill=color, outline=color, width=0)
    canvas.create_arc(x1, y1, x1 + 2 * radio_esquinas, y1 + 2 * radio_esquinas, start=90, extent=90, fill=color, outline=color)
    canvas.create_arc(x2 - 2 * radio_esquinas, y1, x2, y1 + 2 * radio_esquinas, start=0, extent=90, fill=color, outline=color)
    canvas.create_arc(x1, y2 - 2 * radio_esquinas, x1 + 2 * radio_esquinas, y2, start=180, extent=90, fill=color, outline=color)
    canvas.create_arc(x2 - 2 * radio_esquinas, y2 - 2 * radio_esquinas, x2, y2, start=270, extent=90, fill=color, outline=color)
#%% ====== Button -- Image ====================================================
def Button_Image(image_name, lw, lh, container,bg_color,rx,ry,command_function):
    imagen = Image.open(os.path.join(os.getcwd(),"icon") + image_name)
    imagen = imagen.resize((lw,lh), Image.LANCZOS)
    imagen = ImageTk.PhotoImage(imagen)
    button = tk.Button(container, image=imagen, bd=0, bg=bg_color, command=command_function)
    button.image = imagen
    button.place(relx=rx, rely=ry, anchor=tk.CENTER)
    return button
#%% ====== Button -- Image/lambda =============================================
def Button_Image_lambda(image_name, lw, lh, container,bg_color,rx,ry,command_function,label):
    imagen = Image.open(os.path.join(os.getcwd(),"icon") + image_name)
    imagen = imagen.resize((lw,lh), Image.LANCZOS)
    imagen = ImageTk.PhotoImage(imagen)
    button = tk.Button(container, image=imagen, bd=0, bg=bg_color, command=lambda:command_function(label))
    button.image = imagen
    button.place(relx=rx, rely=ry, anchor=tk.CENTER)
    return button 

def Button_Image_lambda3(image_name, lw, lh, container,bg_color,rx,ry,command_function,label1,label2,label3):
    imagen = Image.open(os.path.join(os.getcwd(),"icon") + image_name)
    imagen = imagen.resize((lw,lh), Image.LANCZOS)
    imagen = ImageTk.PhotoImage(imagen)
    button = tk.Button(container, image=imagen, bd=0, bg=bg_color, command=lambda:command_function(label1,label2,label3))
    button.image = imagen
    button.place(relx=rx, rely=ry, anchor=tk.CENTER)
    return button 
#%% ====== Funcion tick modificados ===========================================
def format_tick(value,position):
    value_in_tik = int(value/1000)                                     
    return f'{value_in_tik}k'
#%% ====== Funcion -- Calibrar ================================================
def canva_CLB(datos, canvas_master, relex, reley):
    # datos: Diccionario de datos
    # canvas_master: Frame en donde estara ubicado el grafico
    # ---- Parametros de la grafica -------------------------------------------
    plt.rcParams['savefig.bbox'] = 'tight'
    plt.style.use('default')
    plt.rcParams['font.family'] = 'Calibri'
    fig, ax1 = plt.subplots(figsize=(6, 3))
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.15, right=0.85, top=0.90, bottom=0.20)
    ax1.grid(True, lw=0.3, which='both')
    ax1.tick_params(labelsize=11, width=4)
    
    # ---- Procesar datos -----------------------------------------------------
    df_crc = pd.DataFrame(datos)
    
    # ---- Ordenar datos df si no estan ordenados -----------------------------
    df = df_crc.sort_values(by='Num_Sim', ascending=True)                       # Organiza de menor a mayor
    df = df.reset_index(drop=True)                                              # Reset el index del dataframe
    
    # ---- Configuración del gráfico ------------------------------------------   
    # Configuración de ticks específicos en el eje x
    ax1.xaxis.set_major_formatter(ticker.FuncFormatter(format_tick))
    ax1.set_xticks(df['Num_Sim'])  # Establecer explícitamente las posiciones de los ticks
    ax1.plot(df['Num_Sim'], df['loss'], 'o', color='#262626', markersize=4.5)
    
    ax1.set_xlabel('Número de ses_per_logic_tree', fontsize=12,fontweight='bold')
    
    if df['Event_Based'][0] == "event_based_risk":
        titulo_ylabel = 'Pérdida anual promedio [COP M$]'
    elif df['Event_Based'][0] == "event_based_damage":
        titulo_ylabel = 'Daños estructurales [No. Edificaciones]'
    
    ax1.set_ylabel(titulo_ylabel, fontsize=12, color='#262626',fontweight='bold')
    ax1.tick_params(axis='y', labelcolor='#262626', labelsize=10)
    
    # ---- Añadir error -------------------------------------------------------
    ax2 = ax1.twinx()
    ax2.set_ylabel('Error [%]', color='#C34545',fontweight='bold', fontsize=12)
    ax2.plot(df['Num_Sim'][1:], df['error'][1:], '*', color='#C34545')
    
    # ---- Eje secundario para numero de eventos ------------------------------
    ax3 = ax1.twiny()
    # Los límites del eje x secundario deben coincidir con ax1
    ax3.set_xticks(df['Num_Sim_Ev'])  # Establece los mismos ticks que ax1
    ax3.plot(df['Num_Sim_Ev'], df['loss'], 'o', color='#262626', markersize=4.5)
    ax3.tick_params(axis='x', labelcolor='#262626', labelsize=9)
    ax3.set_xlabel("Número de eventos", fontsize=12,fontweight='bold')  
    
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(fig, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False)
    canvas.get_tk_widget().place(relx=relex, rely=reley, anchor=tk.CENTER, width=600, height=430)
    return canvas

def canva_CLB_Mnz(datos, canvas_master, title_text, relex, reley):
    # datos: Diccionario de datos
    # canvas_master: Frame en donde estara ubicado el grafico
    # ---- Parametros de la grafica -------------------------------------------
    plt.rcParams['savefig.bbox'] = 'tight'
    plt.style.use('default')
    plt.rcParams['font.family'] = 'Calibri'
    fig, ax1 = plt.subplots(figsize=(6, 3))
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.15, right=0.85, top=0.80, bottom=0.10)
    ax1.grid(True, lw=0.3, which='both')
    ax1.tick_params(labelsize=11, width=4)
    
    # ---- Procesar datos -----------------------------------------------------
    df_crc = pd.DataFrame(datos)
    
    # ---- Ordenar datos df si no estan ordenados -----------------------------
    df = df_crc.sort_values(by='Num_Sim', ascending=True)                       # Organiza de menor a mayor
    df = df.reset_index(drop=True)                                              # Reset el index del dataframe
    
    # ---- Configuración del gráfico ------------------------------------------   
    # Configuración de ticks específicos en el eje x
    ax1.xaxis.set_major_formatter(ticker.FuncFormatter(format_tick))
    ax1.set_xticks(df['Num_Sim'])  # Establecer explícitamente las posiciones de los ticks
    ax1.plot(df['Num_Sim'], df['loss'], 'o', color='#262626', markersize=4.5)
    
    ax1.set_xlabel('Número de ses_per_logic_tree', fontsize=12,fontweight='bold')
    
    if df['Event_Based'][0] == "event_based_risk":
        titulo_ylabel = 'Pérdida anual promedio [%]'
    elif df['Event_Based'][0] == "event_based_damage":
        titulo_ylabel = 'Daños estructurales [%]'
    
    ax1.set_ylabel(titulo_ylabel, fontsize=12, color='#262626',fontweight='bold')
    ax1.set_title(title_text,fontsize=11, color='#262626')
    ax1.tick_params(axis='y', labelcolor='#262626', labelsize=10)
    
    # ---- Añadir error -------------------------------------------------------
    ax2 = ax1.twinx()
    ax2.set_ylabel('Error [%]', color='#C34545',fontweight='bold', fontsize=12)
    ax2.plot(df['Num_Sim'][1:], df['error'][1:], '*', color='#C34545')
    
    # ---- Eje secundario para numero de eventos ------------------------------
    ax3 = ax1.twiny()
    # Los límites del eje x secundario deben coincidir con ax1
    ax3.set_xticks(df['Num_Sim_Ev'])  # Establece los mismos ticks que ax1
    ax3.plot(df['Num_Sim_Ev'], df['loss'], 'o', color='#262626', markersize=4.5)
    ax3.tick_params(axis='x', labelcolor='#262626', labelsize=9)
    ax3.set_xlabel("Número de eventos", fontsize=12,fontweight='bold')  
    
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(fig, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False)
    canvas.get_tk_widget().place(relx=relex, rely=reley, anchor=tk.CENTER, width=600, height=430)
    return canvas
#%% ====== Funcion -- Dispersion ==============================================
def canva_DSP(datos, canvas_master, relex, reley):
    # datos: Diccionario de datos
    # canvas_master: Frame en donde estara ubicado el grafico
    # ---- Parametros de la grafica -------------------------------------------
    plt.rcParams['savefig.bbox'] = 'tight'
    plt.style.use('default')
    plt.rcParams['font.family'] = 'Calibri'
    fig, ax1 = plt.subplots(figsize=(6, 3))
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.15, right=0.85, top=0.90, bottom=0.20)
    ax1.grid(True, lw=0.3, which='both')
    ax1.tick_params(labelsize=11, width=4)
    # ---- Procesar datos -----------------------------------------------------
    df_crc = pd.DataFrame(datos)
    
    # ---- Ordenar datos df si no estan ordenados -----------------------------
    df = df_crc.sort_values(by='Num_Sim', ascending=True)                       # Organiza de menor a mayor
    df = df.reset_index(drop=True)                                              # Reset el index del dataframe
    
    # ---- Configuración del gráfico ------------------------------------------   
    # Configuración de ticks específicos en el eje x
    ax1.xaxis.set_major_formatter(ticker.FuncFormatter(format_tick))
    ax1.set_xticks(df['Num_Sim'])  # Establecer explícitamente las posiciones de los ticks
    ax1.plot(df['Num_Sim'], df['loss'], 'o--', markersize=4.5, color='#262626',linewidth=1.0)
    
    ax1.set_xlabel('Número de ses_per_logic_tree', fontsize=12,fontweight='bold')
    
    if df['Event_Based'][0] == "event_based_risk":
        titulo_ylabel = 'Dispersión de las pérdidas anuales promedio'
    elif df['Event_Based'][0] == "event_based_damage":
        titulo_ylabel = 'Dispersión de los daños estructurales anuales promedio'
    
    ax1.set_ylabel(titulo_ylabel, fontsize=10, color='#262626',fontweight='bold')
    ax1.tick_params(axis='y', labelcolor='#262626', labelsize=10)
    
    # ---- Eje secundario para numero de eventos ------------------------------
    ax3 = ax1.twiny()
    # Los límites del eje x secundario deben coincidir con ax1
    ax3.set_xticks(df['Num_Sim_Ev'])  # Establece los mismos ticks que ax1
    ax3.plot(df['Num_Sim_Ev'], df['loss'], 'o', color='#262626', markersize=4.5)
    ax3.tick_params(axis='x', labelcolor='#262626', labelsize=9)
    ax3.set_xlabel("Número de eventos", fontsize=12,fontweight='bold')  
    
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(fig, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False)
    canvas.get_tk_widget().place(relx=relex, rely=reley, anchor=tk.CENTER, width=600, height=430)
    return canvas

def canva_DSP_Mnz(datos, canvas_master, title_text, relex, reley):
    # datos: Diccionario de datos
    # canvas_master: Frame en donde estara ubicado el grafico
    # ---- Parametros de la grafica -------------------------------------------
    plt.rcParams['savefig.bbox'] = 'tight'
    plt.style.use('default')
    plt.rcParams['font.family'] = 'Calibri'
    fig, ax1 = plt.subplots(figsize=(6, 3))
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.15, right=0.85, top=0.80, bottom=0.10)
    ax1.grid(True, lw=0.3, which='both')
    ax1.tick_params(labelsize=11, width=4)
    # ---- Procesar datos -----------------------------------------------------
    df_crc = pd.DataFrame(datos)
    
    # ---- Ordenar datos df si no estan ordenados -----------------------------
    df = df_crc.sort_values(by='Num_Sim', ascending=True)                       # Organiza de menor a mayor
    df = df.reset_index(drop=True)                                              # Reset el index del dataframe
    
    # ---- Configuración del gráfico ------------------------------------------   
    # Configuración de ticks específicos en el eje x
    ax1.xaxis.set_major_formatter(ticker.FuncFormatter(format_tick))
    ax1.set_xticks(df['Num_Sim'])  # Establecer explícitamente las posiciones de los ticks
    ax1.plot(df['Num_Sim'], df['loss'], 'o--', markersize=4.5, color='#262626',linewidth=1.0)
    
    ax1.set_xlabel('Número de ses_per_logic_tree', fontsize=12,fontweight='bold')
    
    if df['Event_Based'][0] == "event_based_risk":
        titulo_ylabel = 'Dispersión de las pérdidas anuales promedio'
    elif df['Event_Based'][0] == "event_based_damage":
        titulo_ylabel = 'Dispersión de los daños estructurales anuales promedio'

    ax1.set_ylabel(titulo_ylabel, fontsize=10, color='#262626',fontweight='bold')
    ax1.set_title(title_text,fontsize=11, color='#262626')
    ax1.tick_params(axis='y', labelcolor='#262626', labelsize=10)
    
    # ---- Eje secundario para numero de eventos ------------------------------
    ax3 = ax1.twiny()
    # Los límites del eje x secundario deben coincidir con ax1
    ax3.set_xticks(df['Num_Sim_Ev'])  # Establece los mismos ticks que ax1
    ax3.plot(df['Num_Sim_Ev'], df['loss'], 'o', color='#262626', markersize=4.5)
    ax3.tick_params(axis='x', labelcolor='#262626', labelsize=9)
    ax3.set_xlabel("Número de eventos", fontsize=12,fontweight='bold')  
    
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(fig, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False)
    canvas.get_tk_widget().place(relx=relex, rely=reley, anchor=tk.CENTER, width=600, height=430)
    return canvas

#%% ====== Funcion -- Perdidas ================================================
def format_tick_EBR(value,position):
    value_in_tik = int(value)                                     
    return f'{value_in_tik}%'

def curva_excedencia(datos, valex_entry):
    # ---- Parametros de la grafica -------------------------------------------
    plt.rcParams['savefig.bbox'] = 'tight'
    plt.style.use('default')
    plt.rcParams['font.family'] = 'Calibri'
    fig = plt.figure(figsize=(6,3))
    fig, ax = plt.subplots()
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.12,right=0.98,top=0.98,bottom=0.11)
    ax.tick_params(labelsize=12, width=4)
    ax.grid(True, lw=0.15, which='both')
    # ---- Configuración del gráfico ------------------------------------------
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(format_tick_EBR))
    ax.plot((datos.loss/(valex_entry*1e6))*100,datos['perdidaTA'],color='#3B3838')
    anoslist = [475,975,2475]
    color = ['#C0504D','#31859C','#BA7296']
    for inx,i in enumerate(anoslist):
        datos['diferencia'] = abs(datos['perdidaTA']-1/i)
        indice_point = datos['diferencia'].idxmin()
        perdidaval = (datos.loc[indice_point,'loss'])*100/(valex_entry*1e6)
        ax.plot([0,perdidaval],[1/i,1/i],'--',color=color[inx])
        ax.plot([perdidaval,perdidaval],[np.min(datos['perdidaTA']),1/i],'--',color=color[inx])
        ax.plot(perdidaval,1/i,'o',color=color[inx],markersize=4.0,label=f'{np.around(perdidaval,1)}% en {i} años')
    ax.set_xlabel('Pérdida anual [% Valor expuesto]',fontsize=12)
    ax.set_ylabel('Tasa de Excedencia [1/año]',fontsize=12)
    ax.set_xlim(np.min((datos.loss/(valex_entry*1e6))*100),np.max((datos.loss/(valex_entry*1e6))*100))
    ax.set_ylim(np.min(datos['perdidaTA']),np.max(datos['perdidaTA']))
    ax.legend()
    ax.set_yscale('log')

    return fig

def canva_crv_EBR(datos, valex_entry, valper_entry, canvas_master, relex, reley):
    # ---- Parametros de la grafica -------------------------------------------
    plt.rcParams['savefig.bbox'] = 'tight'
    plt.style.use('default')
    plt.rcParams['font.family'] = 'Calibri'
    fig = plt.figure(figsize=(6,3))
    fig, ax = plt.subplots()
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.12,right=0.98,top=0.98,bottom=0.11)
    ax.tick_params(labelsize=12, width=4)
    ax.grid(True, lw=0.15, which='both')
    # ---- Configuración del gráfico ------------------------------------------
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(format_tick_EBR))
    ax.plot((datos.loss/(valex_entry*1e6))*100,datos['perdidaTA'],color='#3B3838')
    if valper_entry is not None:
        anoslist = [475,975,2475,valper_entry]
        # color = ['#8064A2','#8CA448','#C0504D','#31859C','#BA7296','#B27E5E']
        color = ['#C0504D','#31859C','#BA7296','#B27E5E']
    else:
        anoslist = [475,975,2475]
        # anoslist = [225,475,975]
        # color = ['#8064A2','#8CA448','#C0504D','#31859C','#BA7296']
        color = ['#C0504D','#31859C','#BA7296']
    for inx,i in enumerate(anoslist):
        datos['diferencia'] = abs(datos['perdidaTA']-1/i)
        indice_point = datos['diferencia'].idxmin()
        perdidaval = (datos.loc[indice_point,'loss'])*100/(valex_entry*1e6)
        ax.plot([0,perdidaval],[1/i,1/i],'--',color=color[inx])
        ax.plot([perdidaval,perdidaval],[np.min(datos['perdidaTA']),1/i],'--',color=color[inx])
        ax.plot(perdidaval,1/i,'o',color=color[inx],markersize=4.0,label=f'{np.around(perdidaval,1)}% en {i} años')
    ax.set_xlabel('Pérdida anual [% Valor expuesto]',fontsize=12)
    ax.set_ylabel('Tasa de Excedencia [1/año]',fontsize=12)
    ax.set_xlim(np.min((datos.loss/(valex_entry*1e6))*100),np.max((datos.loss/(valex_entry*1e6))*100))
    ax.set_ylim(np.min(datos['perdidaTA']),np.max(datos['perdidaTA']))
    ax.legend()
    ax.set_yscale('log')
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(fig, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH,expand=False)
    canvas.get_tk_widget().place(relx=relex,rely=reley,anchor=tk.CENTER, width=550, height=400)
    return canvas

def canva_EBR_taxo(df_expotax,canvas_master,relex,reley):
    # Crear la figura y los ejes
    fig, ax1 = plt.subplots(figsize=(9,7))
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.10, right=0.90, top=0.97, bottom=0.20) #0.27
    # Crear el primer gráfico de barras
    ax1.bar(np.array(range(len(df_expotax.taxonomy))) - 0.2, np.around(df_expotax.valex/1e6,3), width=0.4, label='Valor expuesto [COP Billones]', color='orange',alpha=0.5)
    ax1.set_ylabel('Valor expuesto [COP billones]', color='orange',fontsize=12)
    ax1.tick_params(axis='y', labelcolor='orange',labelsize=10)
    # Crear un segundo eje para el segundo gráfico de barras
    ax2 = ax1.twinx()
    ax2.bar(np.array(range(len(df_expotax.taxonomy))) + 0.2, np.around(df_expotax.loss,3)*0.001, width=0.4, label='Pérdida anual esperada [COP miles de millón]', color='blue',alpha=0.5)
    ax2.set_ylabel('Pérdida anual esperada [COP miles de millón]', color='blue',fontsize=12)
    ax2.tick_params(axis='y', labelcolor='blue',labelsize=10)
    # Añadir la línea de puntos con valores sobre cada punto
    line_data = np.around((df_expotax.loss/df_expotax.valex)*1000,3)
    for i in range(int(len(df_expotax.taxonomy)/2)):
        ax2.plot(np.array(range(len(df_expotax.taxonomy)))[i], line_data[i], 'ko',markersize=3.0)
        ax2.text(np.array(range(len(df_expotax.taxonomy)))[i]+0.4, line_data[i]-0.0015, f'{line_data[i]}‰', color='k', ha='center', fontsize=8)
    for i in range(int(len(df_expotax.taxonomy)/2),len(df_expotax.taxonomy)):
        ax2.plot(np.array(range(len(df_expotax.taxonomy)))[i], line_data[i], 'ko',markersize=3.0)
        ax2.text(np.array(range(len(df_expotax.taxonomy)))[i]+0.4, line_data[i]-0.002, f'{line_data[i]}‰', color='k', ha='center', fontsize=8)
    line, = ax2.plot(np.NaN, np.NaN, 'ko-', markersize=3.0, label='Pérdida anual esperada [‰]')
    ax2.plot(np.array(range(len(df_expotax.taxonomy))),line_data,'k-',linewidth=1, alpha = 0.6)
    x_labels = df_expotax.taxonomy
    # Cambiar los ticks del eje x a letras
    ax1.set_xticks(np.array(range(len(df_expotax.taxonomy))))
    ax1.set_xticklabels(x_labels,rotation=90,fontsize=9)
    # Añadir leyendas y título
    ax1.set_xlabel('Taxonomías',fontsize=13)
    # fig.legend(loc='lower center', bbox_to_anchor=(0.5, -0.003), ncol=2, fontsize=8)
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(fig, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH,expand=False)
    canvas.get_tk_widget().place(relx=relex,rely=reley,anchor=tk.CENTER, width=720, height=500)
    return canvas

def canva_mapPAE(COD_mun,CP_Name,manzana_shp,seccion_shp,area_shp,map_data,Modelo_Expo,canvas_master,relex,reley,Separa_x,Separa_y,User_min_lon,User_max_lon,User_min_lat,User_max_lat):
    
    # =================== VARIABLES MODIFICAR - area_construida ===================
    Title_FontSize = 6.5         # Tamano de letra del nombre de la variable presentada en el mapa
    Cbar_Scale = 0.003           # Espaciado de los ticks del cbar
    Cbar_LabelSize = 4           # Tamano de letra de los ticks del cbar
    Cbar_FontSize = 6.5          # Tamano de letra del titulo del cbar
    Cientific = 0                # Notacion cientifica 1: si 0: no
    # =============================================================================
    
    variable = 'pae_mnz_cop'
    Cmap_text = 'PAE [COP Millones]'
    title1 = 'Pérdida anual esperada en millones'
    title2 = 'de pesos COP por manzana censal'
    viridis_inverted = cm.get_cmap('viridis').reversed()

    
    # Generar dataframe data_figura38
    # Para eso necesito:
        # 1. Modelo de exposicion del municipio (informar en guia el nombre de las columnas)
        
    # Codigo de las manzanas corregido

    cod_mnzdef_model = []
    for mnz in Modelo_Expo.cod_mnz:
        if mnz[0].isalpha():
            mnz2 = str(mnz)
            cod_mnzdef_model.append(mnz2[1::])   
        else:
            mnz2 = str(mnz)
            cod_mnzdef_model.append(mnz2)  
            
    Modelo_Expo.cod_mnz = cod_mnzdef_model

    cod_sccdef_model = []
    for scc in Modelo_Expo.cod_secc:
        if scc[0].isalpha():
            cod_sccdef_model.append(scc[1::])
        else:
            cod_sccdef_model.append(scc)
               
    Modelo_Expo.cod_secc = cod_sccdef_model
    
    # Obtener taxonomia representativa
    Expo_groupby_txn = Modelo_Expo.groupby('tipologia')['area_cons'] 
    Expo_groupby_mnz = Modelo_Expo.groupby('cod_mnz')['area_cons'].sum()

    Area_constr_txn = Expo_groupby_txn.sum()
    df_area = pd.DataFrame({'tipologia':Area_constr_txn.index,'area':Area_constr_txn})
    Txn_Rep = df_area.loc[df_area.area == df_area.area.max()].values[0][0]          
    # Obtener el % de area construida de las manzanas con esa taxonomia representativa
    df_mnz_txn = Modelo_Expo.loc[Modelo_Expo.tipologia == Txn_Rep]
    df_mnz_groupbymnz = df_mnz_txn.groupby('cod_mnz')['area_cons'].sum()
    df_final = pd.DataFrame({'cod_mnz2':df_mnz_groupbymnz.index,'area':df_mnz_groupbymnz})
    df_final = df_final.merge(Expo_groupby_mnz, left_on='cod_mnz2', right_on='cod_mnz', how='left')
    df_final['area_cons'] = (df_final.area/df_final.area_cons)*100 


    # Geopandas para definir los limites del mapa
    map_data_limits = df_final.merge(manzana_shp, left_on='cod_mnz2', right_on='COD_DANE', how='left')
    map_data_limits_df = map_data_limits.to_csv("Map_Limits.csv",index=False)

    
    # Generar mapa de manzanas
    figura = mapa_gen_mnz(Cientific,Separa_x,Separa_y,COD_mun,CP_Name,manzana_shp,seccion_shp,area_shp,map_data,map_data_limits_df,variable,viridis_inverted,Cmap_text,title1,title2,Title_FontSize,Cbar_Scale,Cbar_LabelSize,Cbar_FontSize,User_min_lon,User_max_lon,User_min_lat,User_max_lat)
    
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(figura, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False)
    canvas.get_tk_widget().place(relx=relex, rely=reley, anchor=tk.CENTER, width=972, height=550)
    
    return canvas
    
def canva_mapPAE_Valex(COD_mun,CP_Name,manzana_shp,seccion_shp,area_shp,map_data,Modelo_Expo,canvas_master,relex,reley,Separa_x,Separa_y,User_min_lon,User_max_lon,User_min_lat,User_max_lat):
    
    # =================== VARIABLES MODIFICAR - area_construida ===================
    Title_FontSize = 6.5         # Tamano de letra del nombre de la variable presentada en el mapa
    Cbar_Scale = 0.003           # Espaciado de los ticks del cbar
    Cbar_LabelSize = 4           # Tamano de letra de los ticks del cbar
    Cbar_FontSize = 6.5          # Tamano de letra del titulo del cbar
    Cientific = 0                # Notacion cientifica 1: si 0: no
    # =============================================================================
    
    variable = 'pae_mnz_prc'
    Cmap_text = 'PAE [‰]'
    title1 = 'Pérdida anual esperada en ‰'
    title2 = '(PAE/Valor espuesto) por manzana censal'
    viridis_inverted = cm.get_cmap('viridis').reversed()

    
    # Generar dataframe data_figura38
    # Para eso necesito:
        # 1. Modelo de exposicion del municipio (informar en guia el nombre de las columnas)
        
    # Codigo de las manzanas corregido

    cod_mnzdef_model = []
    for mnz in Modelo_Expo.cod_mnz:
        if mnz[0].isalpha():
            mnz2 = str(mnz)
            cod_mnzdef_model.append(mnz2[1::])   
        else:
            mnz2 = str(mnz)
            cod_mnzdef_model.append(mnz2)  
            
    Modelo_Expo.cod_mnz = cod_mnzdef_model

    cod_sccdef_model = []
    for scc in Modelo_Expo.cod_secc:
        if scc[0].isalpha():
            cod_sccdef_model.append(scc[1::])
        else:
            cod_sccdef_model.append(scc)
               
    Modelo_Expo.cod_secc = cod_sccdef_model
    
    # Obtener taxonomia representativa
    Expo_groupby_txn = Modelo_Expo.groupby('tipologia')['area_cons'] 
    Expo_groupby_mnz = Modelo_Expo.groupby('cod_mnz')['area_cons'].sum()

    Area_constr_txn = Expo_groupby_txn.sum()
    df_area = pd.DataFrame({'tipologia':Area_constr_txn.index,'area':Area_constr_txn})
    Txn_Rep = df_area.loc[df_area.area == df_area.area.max()].values[0][0]          
    # Obtener el % de area construida de las manzanas con esa taxonomia representativa
    df_mnz_txn = Modelo_Expo.loc[Modelo_Expo.tipologia == Txn_Rep]
    df_mnz_groupbymnz = df_mnz_txn.groupby('cod_mnz')['area_cons'].sum()
    df_final = pd.DataFrame({'cod_mnz2':df_mnz_groupbymnz.index,'area':df_mnz_groupbymnz})
    df_final = df_final.merge(Expo_groupby_mnz, left_on='cod_mnz2', right_on='cod_mnz', how='left')
    df_final['area_cons'] = (df_final.area/df_final.area_cons)*100 


    # Geopandas para definir los limites del mapa
    map_data_limits = df_final.merge(manzana_shp, left_on='cod_mnz2', right_on='COD_DANE', how='left')
    map_data_limits_df = map_data_limits.to_csv("Map_Limits.csv",index=False)

    
    # Generar mapa de manzanas
    figura = mapa_gen_mnz(Cientific,Separa_x,Separa_y,COD_mun,CP_Name,manzana_shp,seccion_shp,area_shp,map_data,map_data_limits_df,variable,viridis_inverted,Cmap_text,title1,title2,Title_FontSize,Cbar_Scale,Cbar_LabelSize,Cbar_FontSize,User_min_lon,User_max_lon,User_min_lat,User_max_lat)
    
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(figura, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False)
    canvas.get_tk_widget().place(relx=relex, rely=reley, anchor=tk.CENTER, width=972, height=550)
    
    return canvas


def canva_mapPAE_secc(COD_mun,CP_Name,seccion_shp,area_shp,map_data,Modelo_Expo,canvas_master,relex,reley,Separa_x,Separa_y,User_min_lon,User_max_lon,User_min_lat,User_max_lat):
    
    # =================== VARIABLES MODIFICAR - area_construida ===================
    Title_FontSize = 6.5         # Tamano de letra del nombre de la variable presentada en el mapa
    Cbar_Scale = 0.003           # Espaciado de los ticks del cbar
    Cbar_LabelSize = 4           # Tamano de letra de los ticks del cbar
    Cbar_FontSize = 6.5          # Tamano de letra del titulo del cbar
    Cientific = 0                # Notacion cientifica 1: si 0: no
    # =============================================================================
    
    variable = 'pae_scc_cop'
    Cmap_text = 'PAE [COP Millones]'
    title1 = 'Pérdida anual esperada en millones'
    title2 = 'de pesos COP por sección urbana'
    viridis_inverted = cm.get_cmap('viridis').reversed()
    
    # Generar dataframe data_figura38
    # Para eso necesito:
        # 1. Modelo de exposicion del municipio (informar en guia el nombre de las columnas)
        
    # Codigo de las manzanas corregido

    cod_mnzdef_model = []
    for mnz in Modelo_Expo.cod_mnz:
        if mnz[0].isalpha():
            mnz2 = str(mnz)
            cod_mnzdef_model.append(mnz2[1::])   
        else:
            mnz2 = str(mnz)
            cod_mnzdef_model.append(mnz2)  
            
    Modelo_Expo.cod_mnz = cod_mnzdef_model

    cod_sccdef_model = []
    for scc in Modelo_Expo.cod_secc:
        if scc[0].isalpha():
            cod_sccdef_model.append(scc[1::])
        else:
            cod_sccdef_model.append(scc)
               
    Modelo_Expo.cod_secc = cod_sccdef_model
    
    # Obtener taxonomia representativa
    Expo_groupby_txn = Modelo_Expo.groupby('tipologia')['area_cons'] 
    Expo_groupby_mnz = Modelo_Expo.groupby('cod_secc')['area_cons'].sum()

    Area_constr_txn = Expo_groupby_txn.sum()
    df_area = pd.DataFrame({'tipologia':Area_constr_txn.index,'area':Area_constr_txn})
    Txn_Rep = df_area.loc[df_area.area == df_area.area.max()].values[0][0]          
    # Obtener el % de area construida de las manzanas con esa taxonomia representativa
    df_mnz_txn = Modelo_Expo.loc[Modelo_Expo.tipologia == Txn_Rep]
    df_mnz_groupbymnz = df_mnz_txn.groupby('cod_secc')['area_cons'].sum()
    df_final = pd.DataFrame({'cod_scc2':df_mnz_groupbymnz.index,'area':df_mnz_groupbymnz})
    df_final = df_final.merge(Expo_groupby_mnz, left_on='cod_scc2', right_on='cod_secc', how='left')
    df_final['area_cons'] = (df_final.area/df_final.area_cons)*100 


    # Geopandas para definir los limites del mapa
    map_data_limits = df_final.merge(seccion_shp, left_on='cod_scc2', right_on='COD_SECC', how='left')
    map_data_limits_df = map_data_limits.to_csv("Map_Limits_secc.csv",index=False)

    
    # Generar mapa de manzanas
    figura = mapa_gen_scc(Cientific,Separa_x,Separa_y,COD_mun,CP_Name,seccion_shp,area_shp,map_data,map_data_limits_df,variable,viridis_inverted,Cmap_text,title1,title2,Title_FontSize,Cbar_Scale,Cbar_LabelSize,Cbar_FontSize,User_min_lon,User_max_lon,User_min_lat,User_max_lat)
    
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(figura, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False)
    canvas.get_tk_widget().place(relx=relex, rely=reley, anchor=tk.CENTER, width=972, height=550)
    
    return canvas

def canva_mapPAEValex_secc(COD_mun,CP_Name,seccion_shp,area_shp,map_data,Modelo_Expo,canvas_master,relex,reley,Separa_x,Separa_y,User_min_lon,User_max_lon,User_min_lat,User_max_lat):
    
    # =================== VARIABLES MODIFICAR - area_construida ===================
    Title_FontSize = 6.5         # Tamano de letra del nombre de la variable presentada en el mapa
    Cbar_Scale = 0.003           # Espaciado de los ticks del cbar
    Cbar_LabelSize = 4           # Tamano de letra de los ticks del cbar
    Cbar_FontSize = 6.5          # Tamano de letra del titulo del cbar
    Cientific = 0                # Notacion cientifica 1: si 0: no
    # =============================================================================
    
    variable = 'pae_scc_prc'
    Cmap_text = 'PAE [‰]'
    title1 = 'Pérdida anual esperada en ‰'
    title2 = '(PAE/Valor espuesto) por sección urbana'
    viridis_inverted = cm.get_cmap('viridis').reversed()
    
    # Generar dataframe data_figura38
    # Para eso necesito:
        # 1. Modelo de exposicion del municipio (informar en guia el nombre de las columnas)
        
    # Codigo de las manzanas corregido

    cod_mnzdef_model = []
    for mnz in Modelo_Expo.cod_mnz:
        if mnz[0].isalpha():
            mnz2 = str(mnz)
            cod_mnzdef_model.append(mnz2[1::])   
        else:
            mnz2 = str(mnz)
            cod_mnzdef_model.append(mnz2)  
            
    Modelo_Expo.cod_mnz = cod_mnzdef_model

    cod_sccdef_model = []
    for scc in Modelo_Expo.cod_secc:
        if scc[0].isalpha():
            cod_sccdef_model.append(scc[1::])
        else:
            cod_sccdef_model.append(scc)
               
    Modelo_Expo.cod_secc = cod_sccdef_model
    
    # Obtener taxonomia representativa
    Expo_groupby_txn = Modelo_Expo.groupby('tipologia')['area_cons'] 
    Expo_groupby_mnz = Modelo_Expo.groupby('cod_secc')['area_cons'].sum()

    Area_constr_txn = Expo_groupby_txn.sum()
    df_area = pd.DataFrame({'tipologia':Area_constr_txn.index,'area':Area_constr_txn})
    Txn_Rep = df_area.loc[df_area.area == df_area.area.max()].values[0][0]          
    # Obtener el % de area construida de las manzanas con esa taxonomia representativa
    df_mnz_txn = Modelo_Expo.loc[Modelo_Expo.tipologia == Txn_Rep]
    df_mnz_groupbymnz = df_mnz_txn.groupby('cod_secc')['area_cons'].sum()
    df_final = pd.DataFrame({'cod_scc2':df_mnz_groupbymnz.index,'area':df_mnz_groupbymnz})
    df_final = df_final.merge(Expo_groupby_mnz, left_on='cod_scc2', right_on='cod_secc', how='left')
    df_final['area_cons'] = (df_final.area/df_final.area_cons)*100 


    # Geopandas para definir los limites del mapa
    map_data_limits = df_final.merge(seccion_shp, left_on='cod_scc2', right_on='COD_SECC', how='left')
    map_data_limits_df = map_data_limits.to_csv("Map_Limits_secc.csv",index=False)

    
    # Generar mapa de manzanas
    figura = mapa_gen_scc(Cientific,Separa_x,Separa_y,COD_mun,CP_Name,seccion_shp,area_shp,map_data,map_data_limits_df,variable,viridis_inverted,Cmap_text,title1,title2,Title_FontSize,Cbar_Scale,Cbar_LabelSize,Cbar_FontSize,User_min_lon,User_max_lon,User_min_lat,User_max_lat)
    
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(figura, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False)
    canvas.get_tk_widget().place(relx=relex, rely=reley, anchor=tk.CENTER, width=972, height=550)
    
    return canvas

def Gen_Tabla_Resume_PRD(valorexp,PAE_mill,PE_mill):
    
    # Crear un nuevo libro de Excel y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    # Configuracion de la primera linea 
    ws.merge_cells('B2:C2')
    ws['B2'] = 'Valor expuesto'
    ws['D2'] = '[COP millones]'
    ws['E2'] = valorexp
    # Configuracion de la segunda linea 
    ws.merge_cells('B3:C4')
    ws['B3'] = 'Pérdida anual esperada'
    ws['D3'] = '[COP millones]'
    ws['E3'] = PAE_mill
    ws['D4'] = '[‰]'
    ws['E4'] = (PAE_mill/valorexp)*1000
    # Configuracion de la tercera linea
    ws.merge_cells('B5:E5')
    ws['B5'] = 'Pérdida máxima probable'
    # Configuracion de la tercera linea 
    ws.merge_cells('B6:B7')
    ws['B6'] = 'Periodo de retorno'
    ws.merge_cells('C6:C7')
    ws['C6'] = 'Probabilidad de excedencia en 50 años'
    ws.merge_cells('D6:E7')
    ws['D6'] = 'Pérdida esperada'
    # Configuracion de la cuarta linea
    ws['B8'] = '[años]'
    ws['C8'] = '[%]'
    ws['D8'] = '[COP millones]'
    ws['E8'] = '[%]'
    # Configuracion de la quinta linea
    ws['B9'] = 31
    ws['C9'] = (1-np.exp(-50/31))*100
    ws['D9'] = PE_mill[0]
    ws['E9'] = (PE_mill[0]/valorexp)*100
    # Configuracion de la sexta linea
    ws['B10'] = 225
    ws['C10'] = (1-np.exp(-50/225))*100
    ws['D10'] = PE_mill[1]
    ws['E10'] = (PE_mill[1]/valorexp)*100
    # Configuracion de la septima linea
    ws['B11'] = 475
    ws['C11'] = (1-np.exp(-50/475))*100
    ws['D11'] = PE_mill[2]
    ws['E11'] = (PE_mill[2]/valorexp)*100
    # Configuracion de la octava linea
    ws['B12'] = 975
    ws['C12'] = (1-np.exp(-50/975))*100
    ws['D12'] = PE_mill[3]
    ws['E12'] = (PE_mill[3]/valorexp)*100
    # Configuracion de la novena linea
    ws['B13'] = 1475
    ws['C13'] = (1-np.exp(-50/1475))*100
    ws['D13'] = PE_mill[4]
    ws['E13'] = (PE_mill[4]/valorexp)*100

    # Alinear el contenido de las celdas al centro
    for row in ws.iter_rows(min_row=2, max_row=13, min_col=2, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    # Configurar el tamaño de las celdas
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16

    # Agregar bordes a todas las celdas
    for row in ws.iter_rows(min_row=2, max_row=13, min_col=2, max_col=5):
        for cell in row:
            cell.border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
            
    Table_Resu = wb.save
    
    return Table_Resu

def Gen_Tabla_taxonomia_PRD(df_expotax):
    
    # Crear un nuevo libro de Excel y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    # Configuracion de la primera fila
    ws.merge_cells('B2:B3')
    ws['B2'] = 'Tipología constructiva'
    ws['C2'] = 'Valor expuesto'
    ws.merge_cells('D2:E2')
    ws['D2'] = 'Pérdida anual esperada'
    # Configuración de la segunda fila
    ws['C3'] = '[COP miles de millón]'
    ws['D3'] = '[COP miles de millón]'
    ws['E3'] = '[‰]'
    # Configuracion taxonomia
    valexpuesto,paecop = [],[]
    for index, txn in enumerate(df_expotax.taxonomy):
        ws['B'+str(index+4)] = txn
        ws['C'+str(index+4)] = np.around(df_expotax.valex[index],3)*0.001
        valexpuesto.append((df_expotax.valex[index])*0.001)
        
        ws['D'+str(index+4)] = np.around(df_expotax.loss[index],3)*0.001
        paecop.append((df_expotax.loss[index])*0.001)
        
        ws['E'+str(index+4)] = str(np.around((df_expotax.loss[index]/df_expotax.valex[index])*1000,3))+'‰'

    ws['B'+str(index+5)] = 'Total'
    ws['C'+str(index+5)] = np.sum(valexpuesto)
    ws['D'+str(index+5)] = np.sum(paecop)
    ws['E'+str(index+5)] = str(np.around((np.sum(df_expotax.loss)/np.sum(df_expotax.valex))*1000,3))+'‰'

    # Alinear el contenido de las celdas al centro
    for row in ws.iter_rows(min_row=2, max_row=len(df_expotax.taxonomy)+4, min_col=2, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    # Configurar el tamaño de las celdas
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 11

    # Agregar bordes a todas las celdas
    for row in ws.iter_rows(min_row=2, max_row=len(df_expotax.taxonomy)+4, min_col=2, max_col=5):
        for cell in row:
            cell.border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
    # Guardar el archivo 
    Table_Resu = wb.save
    
    return Table_Resu

def Diagrama_Taxonomia_PRD(df_expotax):
    
    # Crear la figura y los ejes
    fig, ax1 = plt.subplots(figsize=(9,7))
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.10, right=0.90, top=0.97, bottom=0.20) #0.27
    # Crear el primer gráfico de barras
    ax1.bar(np.array(range(len(df_expotax.taxonomy))) - 0.2, np.around(df_expotax.valex/1e6,3), width=0.4, label='Valor expuesto [COP Billones]', color='orange',alpha=0.5)
    ax1.set_ylabel('Valor expuesto [COP billones]', color='orange',fontsize=12)
    ax1.tick_params(axis='y', labelcolor='orange',labelsize=10)
    # Crear un segundo eje para el segundo gráfico de barras
    ax2 = ax1.twinx()
    ax2.bar(np.array(range(len(df_expotax.taxonomy))) + 0.2, np.around(df_expotax.loss,3)*0.001, width=0.4, label='Pérdida anual esperada [COP miles de millón]', color='blue',alpha=0.5)
    ax2.set_ylabel('Pérdida anual esperada [COP miles de millón]', color='blue',fontsize=12)
    ax2.tick_params(axis='y', labelcolor='blue',labelsize=10)
    # Añadir la línea de puntos con valores sobre cada punto
    line_data = np.around((df_expotax.loss/df_expotax.valex)*1000,3)
    for i in range(int(len(df_expotax.taxonomy)/2)):
        ax2.plot(np.array(range(len(df_expotax.taxonomy)))[i], line_data[i], 'ko',markersize=3.0)
        ax2.text(np.array(range(len(df_expotax.taxonomy)))[i]+0.4, line_data[i]-0.0015, f'{line_data[i]}‰', color='k', ha='center', fontsize=8)
    for i in range(int(len(df_expotax.taxonomy)/2),len(df_expotax.taxonomy)):
        ax2.plot(np.array(range(len(df_expotax.taxonomy)))[i], line_data[i], 'ko',markersize=3.0)
        ax2.text(np.array(range(len(df_expotax.taxonomy)))[i]+0.4, line_data[i]-0.002, f'{line_data[i]}‰', color='k', ha='center', fontsize=8)
    line, = ax2.plot(np.NaN, np.NaN, 'ko-', markersize=3.0, label='Pérdida anual esperada [‰]')
    ax2.plot(np.array(range(len(df_expotax.taxonomy))),line_data,'k-',linewidth=1, alpha = 0.6)
    x_labels = df_expotax.taxonomy
    # Cambiar los ticks del eje x a letras
    ax1.set_xticks(np.array(range(len(df_expotax.taxonomy))))
    ax1.set_xticklabels(x_labels,rotation=90,fontsize=9)
    # Añadir leyendas y título
    ax1.set_xlabel('Taxonomías',fontsize=13)
    
    return fig

#%% ====== Libreria mapas =====================================================

def LenMargin_value(value):
    
    len_margin = np.around(1/(len(value)-1),4)
    
    return len_margin

def scientific_formatter(x, pos):
    return f'{x:.0e}'

def extract_coordinates_safe(polygon_str):
    """
    Extracts the coordinates from a POLYGON string and returns them as a list of tuples (latitude, longitude).
    """
    if not isinstance(polygon_str, str):
        # Si la entrada no es una cadena, la convierte en una
        polygon_str = str(polygon_str)

    # Encuentra todas las coincidencias de patrones de coordenadas en la cadena del polígono
    coords = re.findall(r'(-?\d+\.\d+) (-?\d+\.\d+)', polygon_str)
    # Convierte las coordenadas a flotantes y las guarda en tuplas (longitud, latitud)
    return [(float(lon), float(lat)) for lon, lat in coords]

def unir_valores(lista):
    resultado = []
    for sublista in lista:
        # Convertir cada número a string y luego unirlos
        texto = ''.join(str(num) for num in sublista)
        resultado.append(texto)
    return resultado

def agrupar_pisos(lista_pisos):
    # Grupos definidos
    grupos = [
        (1, [1]),             # Edificios de 1 piso
        (2, [2]),             # Edificios de 2 pisos
        (3, [3, 4]),          # Edificios de 3 y 4 pisos
        (4, range(5, 8)),     # Edificios de 5 a 7 pisos
        (5, range(8, 11))     # Edificios de 8 a 10 pisos
    ]

    # Diccionario para almacenar los grupos
    grupos_pisos = {grupo: [] for grupo, _ in grupos}
    grupos_pisos[6] = []  # Grupo para edificios de más de 10 pisos

    # Asignar cada piso a su grupo correspondiente
    for piso in lista_pisos:
        asignado = False
        for grupo, rango in grupos:
            if piso in rango:
                grupos_pisos[grupo].append(piso)
                asignado = True
                break
        if not asignado:
            grupos_pisos[6].append(piso)  # Añadir a 'más de 10'

    # Crear las listas finales
    lista_pisos_agrupados = [grupos_pisos[grupo] for grupo, _ in grupos if grupos_pisos[grupo]]
    if grupos_pisos[6]:
        lista_pisos_agrupados.append(grupos_pisos[6])
    lista_nombre_pisos = ['un piso', 'dos pisos', '3 y 4 pisos', '5 a 7 pisos', '8 a 10 pisos', '12 y 17 pisos']

    return lista_pisos_agrupados, lista_nombre_pisos

def filtrar_por_pisos(df, pisos):
    """
    Filtra el DataFrame 'df' para incluir solo las filas donde 'n_piso' está en la lista 'pisos'.

    :param df: DataFrame a filtrar.
    :param pisos: Lista de números de pisos a incluir.
    :return: DataFrame filtrado.
    """
    return df[df['n_piso'].isin(pisos)]

#%% generar mapa manzanas
def mapa_gen_mnz(Cientific,Separa_x,Separa_y,COD_mun,CP_Name,manzana_shp,seccion_shp,area_shp,map_data,map_data_limits,variable,viridis_inverted,Cmap_text,title1,title2,Title_FontSize,Cbar_Scale,Cbar_LabelSize,Cbar_FontSize,User_min_lon,User_max_lon,User_min_lat,User_max_lat):
        
    fig = plt.figure(figsize=(10.6, 6))
    fig.set_facecolor('white')
    
    # Los valores son [left, bottom, width, height] en fracciones de la figura
    rect_ax = [0.05, 0.1, 0.65, 0.8]
    ax = fig.add_axes(rect_ax)
    
    # 1). Definir el tamaño de las margenes
    # .... Cargar geopandas para el limite del mapa ...........................
    datadf = pd.read_csv(os.path.join(os.getcwd(),'Map_Limits.csv')) 
    
    # .... Aplicar la función corregida a cada fila de la columna 'geometry' 
    # y crear una lista de todas las coordenadas ..............................
    all_coords_safe = []
    datadf['geometry'].apply(lambda x: all_coords_safe.extend(extract_coordinates_safe(x)))
    # .... Calcular los valores máximos y mínimos de longitud y latitud .......
    max_lat = max(all_coords_safe, key=lambda x: x[1])[1]
    min_lat = min(all_coords_safe, key=lambda x: x[1])[1]
    max_lon = max(all_coords_safe, key=lambda x: x[0])[0]
    min_lon = min(all_coords_safe, key=lambda x: x[0])[0]
    
    # ======================= LINEA MODIFICA EL USUARIO =======================
    max_lon_new = np.ceil((min_lon) * 50) / 50 + User_min_lon
    min_lon_new = np.floor((max_lon) * 50) / 50 + User_max_lon
    
    max_lat_new = np.floor((min_lat) * 50) / 50 + User_min_lat
    min_lat_new = np.ceil((max_lat) * 50) / 50 + User_max_lat
    # =========================================================================
    
    if Separa_x == 1:
        separa_num_x = 0.01
    
    if Separa_x == 2:
        separa_num_x = 0.02
        
    if Separa_y == 1:
        separa_num_y = 0.01
    
    if Separa_y == 2:
        separa_num_y = 0.02
    
    ticks_x = np.arange(min_lon_new, max_lon_new - separa_num_x, - separa_num_x)
    ticks_y = np.arange(min_lat_new, max_lat_new - separa_num_y, - separa_num_y)
    
    # Dependiendo de la separación de longitud o latitud
    len_marginx = LenMargin_value(ticks_x)*Separa_x    # Longitud margen en x
    len_marginy = LenMargin_value(ticks_y)*Separa_y    # Longitud margen en y

    
    
    # 2). Definir posicion de ax_legend
    #Ecuacion_linea =-0.2037647997871495 x + 0.825043235333245
    
    ax_left = np.around(-0.2037647997871495*(len_marginx/len_marginy) + 0.825043235333245,2)+0.01
    
    rect_ax_legend = [ax_left, 0.1, 0.18, 0.8]
    ax_legend = fig.add_axes(rect_ax_legend)
    
    
    # ax -- axis del mapa
    # ax_legend -- axis para presentacion del mapa
    # =========================================================================
    
    # .... Cargar los shape files .............................................
    valor_maximo = map_data[variable].max()
    
    if valor_maximo < 1:
        norm = Normalize(vmin=0, vmax=1)
        map_data.plot(column=variable, ax=ax, edgecolor='grey', alpha=1.0, cmap=viridis_inverted, linewidth=0.4, norm=norm)
    else:
        map_data.plot(column=variable, ax=ax, edgecolor='grey', alpha=1.0, cmap=viridis_inverted, linewidth=0.4)
        
    manzana_shp.plot(ax=ax, edgecolor='grey', facecolor="none", alpha=1.0, linewidth=0.4)
    seccion_shp.plot(ax=ax, edgecolor='black', facecolor="none", alpha=1.0, linewidth=0.5)
    area_shp.plot(ax=ax, edgecolor='black', facecolor="none", alpha=1.0, linewidth=1.0)
    # .... Brujula señalando el norte .........................................
    x, y, arrow_length = 0.95, 0.95, 0.1
    ax.annotate('N', xy=(x, y), xytext=(x, y-arrow_length),
                arrowprops=dict(facecolor='black', width=3, headwidth=9),
                ha='center', va='center', fontsize=11, xycoords=ax.transAxes)
    # .... Cbar - Resultados Riesgo ...........................................
    # Calcula la varianza de los datos para verificar si todos los valores son iguales
    varianza_datos = np.var(map_data[variable])
    if varianza_datos == 0:
        # Todos los valores son iguales
        
        # lista de valores únicos, incluidos los NaN
        valores_unicos = map_data[variable].unique()
        # Filtra los valores, eliminando los NaN
        valores_sin_nan = valores_unicos[~np.isnan(valores_unicos)]
        if valores_sin_nan.size == 1:
            valor_unico = valores_sin_nan[0]
        else:
            # Manejar el caso de múltiples valores o ningún valor después de eliminar NaN
            valor_unico = None  
    
        norm = Normalize(vmin=valor_unico, vmax=valor_unico)
    
        # Ajusta las dimensiones del cax si es necesario
        cax = fig.add_axes([rect_ax_legend[0]+0.05, 0.270, 0.08, 0.015]) # [left, bottom, width, height]
        
        # Obtener el colormap viridis
        viridis = plt.cm.get_cmap('viridis', 256)
        
        # El primer color de viridis invertido es el último color de viridis
        primer_color_viridis_inverted = viridis.colors[-1]
        
        # Crear un colormap personalizado que consiste solo en ese color
        colormap_personalizado = ListedColormap([primer_color_viridis_inverted])
        
        # Crea el colorbar con un solo color
        cbar = fig.colorbar(cm.ScalarMappable(norm=norm, cmap=colormap_personalizado), cax=cax, orientation='horizontal')
        
        # Establece un solo tick con el valor único
        cbar.set_ticks([valor_unico])
        cbar.ax.tick_params(axis='x', labelsize=Cbar_LabelSize)
        
    else:
        if valor_maximo < 1:
            cax = fig.add_axes([rect_ax_legend[0]+0.03, 0.270, 0.115, 0.015]) # Ajusta como sea necesario
            cbar = fig.colorbar(cm.ScalarMappable(norm=norm, cmap=viridis_inverted), cax=cax, orientation='horizontal')
            valores = np.linspace(0, 1, 5)
            
        else:
            norm = Normalize(vmin=np.floor(map_data[variable].min()), vmax=np.ceil(map_data[variable].max()))
        
            cax = fig.add_axes([rect_ax_legend[0]+0.03, 0.270, 0.115, 0.015]) # Ajusta como sea necesario
            cbar = fig.colorbar(cm.ScalarMappable(norm=norm, cmap=viridis_inverted), cax=cax, orientation='horizontal')
        
            valores = np.linspace(np.floor(map_data[variable].min()), np.ceil(map_data[variable].max()), 6)
            
        if Cientific == 1:
            formatter = FuncFormatter(scientific_formatter)
            cbar.set_ticks(np.round(valores[1:-1] / Cbar_Scale) * Cbar_Scale)
            cbar.ax.xaxis.set_major_formatter(formatter)
            cbar.ax.tick_params(axis='x', labelsize=Cbar_LabelSize)
        else:
            cbar.set_ticks(np.round(valores[1:-1] / Cbar_Scale) * Cbar_Scale)
            cbar.ax.tick_params(axis='x', labelsize=Cbar_LabelSize)
    
    ax_legend.text(0.5, 0.25, Cmap_text, transform=ax_legend.transAxes,ha='center', va='center', fontsize=Cbar_FontSize)
    # .... Eliminar los ejes de ax_legend .....................................
    ax_legend.set_axis_off()
    # .... En ax_legend anadir los logos de ACOFI y del servicio .............. 
    # Logo ACOFI
    image_path = os.path.join(os.path.join(os.getcwd(), "icon"),'acofilogo.png')
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.0215)
    ab = offsetbox.AnnotationBbox(imagebox, (0.25, 0.95), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)
    # Logo Servicio Geologico
    image_path = os.path.join(os.path.join(os.getcwd(), "icon"),'serviciogc.png')
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.0515)
    ab = offsetbox.AnnotationBbox(imagebox, (0.75, 0.95), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)
    
    
    # .... Generar recuadros para la leyenda ..................................
    ax_legend.add_patch(Rectangle((0.02, 0.785), 0.96, 0.10, fill=True, edgecolor='grey', facecolor='white', linewidth=1.0, transform=ax_legend.transAxes, clip_on=False))
    ax_legend.text(0.5, 0.85, 'Modelo Nacional', transform=ax_legend.transAxes,ha='center', va='center', fontsize=7.8,fontweight='bold')
    ax_legend.text(0.5, 0.82, 'de Riesgo Sísmico', transform=ax_legend.transAxes,ha='center', va='center', fontsize=7.8,fontweight='bold')
    
    ax_legend.add_patch(Rectangle((0.02, 0.67), 0.96, 0.10, fill=True, edgecolor='grey', facecolor='white', linewidth=1.0, transform=ax_legend.transAxes, clip_on=False))
    ax_legend.text(0.5, 0.75, COD_mun, transform=ax_legend.transAxes,ha='center', va='center', fontsize=6,fontweight='bold')
    ax_legend.text(0.5, 0.71, CP_Name, transform=ax_legend.transAxes,ha='center', va='center', fontsize=9,fontweight='bold')
    
    ax_legend.add_patch(Rectangle((0.02, 0.555), 0.96, 0.10, fill=True, edgecolor='grey', facecolor='white', linewidth=1.0, transform=ax_legend.transAxes, clip_on=False))
    ax_legend.text(0.5, 0.615, title1, transform=ax_legend.transAxes,ha='center', va='center', fontsize=Title_FontSize,fontweight='bold')
    ax_legend.text(0.5, 0.59, title2, transform=ax_legend.transAxes,ha='center', va='center', fontsize=Title_FontSize,fontweight='bold')
    
    ax_legend.add_patch(Rectangle((0.02, 0.1), 0.96, 0.44, fill=True, edgecolor='grey', facecolor='white', linewidth=1.0, transform=ax_legend.transAxes, clip_on=False))
    ax_legend.text(0.235, 0.5, 'Leyenda', transform=ax_legend.transAxes,ha='center', va='center', fontsize=8.5)
    ax_legend.text(0.26, 0.45, 'MGN (DANE)', transform=ax_legend.transAxes,ha='center', va='center', fontsize=6,fontweight='bold')
    
    # ax_legend.add_patch(Rectangle((0.715, 0.43), 0.025, 0.015, edgecolor='black', facecolor='none', transform=fig.transFigure))

    
    ax_legend.add_patch(Rectangle((0.105, 0.405), 0.125, 0.02, fill=True, edgecolor='black', facecolor='white', linewidth=1.4, transform=ax_legend.transAxes, clip_on=False))
    ax_legend.text(0.40, 0.412, 'Área censal', transform=ax_legend.transAxes,ha='center', va='center', fontsize=6)
    ax_legend.add_patch(Rectangle((0.105, 0.368), 0.125, 0.02, fill=True, edgecolor='black', facecolor='white', linewidth=0.9, transform=ax_legend.transAxes, clip_on=False))
    ax_legend.text(0.44, 0.375, 'Sección urbana', transform=ax_legend.transAxes,ha='center', va='center', fontsize=6)
    ax_legend.add_patch(Rectangle((0.105, 0.331), 0.125, 0.02, fill=True, edgecolor='grey', facecolor='white', linewidth=0.9, transform=ax_legend.transAxes, clip_on=False))
    ax_legend.text(0.39, 0.339, 'Manzana', transform=ax_legend.transAxes,ha='center', va='center', fontsize=6)
    
    ax_legend.text(0.34, 0.29, 'Resultados riesgo', transform=ax_legend.transAxes,ha='center', va='center', fontsize=6,fontweight='bold')
    
    image_path = os.path.join(os.path.join(os.getcwd(), "icon"),'logosabana.png')
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.17)
    ab = offsetbox.AnnotationBbox(imagebox, (0.25, 0.05), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)
    
    image_path = os.path.join(os.path.join(os.getcwd(), "icon"),'logomedellin.png')
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.12)
    ab = offsetbox.AnnotationBbox(imagebox, (0.75, 0.05), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)
    
    # .... Añadir mapa base ...................................................
    ctx.add_basemap(ax, crs=map_data.crs.to_string(), source=ctx.providers.CartoDB.Positron, zoom=13, alpha=0.5)
    
    
    ax.set_xlim([np.ceil((min_lon) * 50) / 50 + User_min_lon ,np.floor((max_lon) * 50) / 50 + User_max_lon])
    ax.set_ylim([np.floor((min_lat) * 50) / 50 + User_min_lat , np.ceil((max_lat) * 50) / 50 + User_max_lat])
    
    ax.set_xticks(ticks_x)
    ax.set_yticks(ticks_y)
    

    # ============ LINEA PARA MODIFICAR SEGUN CADA CASO =======================
    # .... Generar margenes del mapa ..........................................
    suma = 0
    for index in range(len(ticks_x)-1):
        if suma%2 == 0:
            rect = Rectangle((len_marginx*suma, 0.0), len_marginx, 0.008, fill=True, edgecolor='black', facecolor='black', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((len_marginx*suma, 1.0-0.008), len_marginx, 0.008, fill=True, edgecolor='black', facecolor='black', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1
        else:
            rect = Rectangle((len_marginx*suma, 0.0), len_marginx, 0.008, fill=True, edgecolor='black', facecolor='white', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((len_marginx*suma, 1.0-0.008), len_marginx, 0.008, fill=True, edgecolor='black', facecolor='white', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1
    
    suma = 0
    for index in range(len(ticks_y)-1):
        if suma%2 == 0:
            rect = Rectangle((0.0, len_marginy*suma), 0.007, len_marginy, fill=True, edgecolor='black', facecolor='black', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((1.0 - 0.007, len_marginy*suma), 0.007, len_marginy, fill=True, edgecolor='black', facecolor='black', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1
        else:
            rect = Rectangle((0.0, len_marginy*suma), 0.007, len_marginy, fill=True, edgecolor='black', facecolor='white', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((1.0 - 0.007, len_marginy*suma), 0.007, len_marginy, fill=True, edgecolor='black', facecolor='white', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1
    # =========================================================================
    
    xloc_scale = 0.58
    yloc_scale = 0.08
    
    # =========================================================================
    
    # .... Generar grilla .....................................................
    ax.grid(True, which='both', color='grey', linewidth=0.8, linestyle='-', alpha=0.2)
    len_marginx = len_marginx-len_marginx/5
    # .... Tamano total de la barra escala ....................................
    # Tamano total de la escala (2km + 0.5km)
    # (len_marginx + len_marginx/4) son 2km --> 0.5km son (len_marginx + len_marginx/4)/4
    margen_total = (len_marginx + len_marginx/4) + (len_marginx + len_marginx/4)/4
    
    if Separa_x == 1:
        ax.text((xloc_scale+0.01)+margen_total, 0.03+yloc_scale, '1km', transform=ax.transAxes,ha='center', va='center', fontsize=6)
        
    if Separa_x == 2:
        ax.text((xloc_scale+0.01)+margen_total, 0.03+yloc_scale, '2km', transform=ax.transAxes,ha='center', va='center', fontsize=6)
    
    rect = Rectangle((xloc_scale, yloc_scale), margen_total, 0.011, fill=True, edgecolor='black', facecolor='white', linewidth=0.8, transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    # Tamano de 1km a 1.5km (son 0.5km para colorear en negro)
    tamano1 = (len_marginx + len_marginx/4)/4
    
    if Separa_x == 1:
        ax.text((xloc_scale+margen_total)-(tamano1) , 0.03+yloc_scale, '.75', transform=ax.transAxes,ha='center', va='center', fontsize=5)
        ax.text((xloc_scale+margen_total)-(tamano1)*2 , 0.03+yloc_scale, '0.5', transform=ax.transAxes,ha='center', va='center', fontsize=5)

    if Separa_x == 2:
        ax.text((xloc_scale+margen_total)-(tamano1) , 0.03+yloc_scale, '1.5', transform=ax.transAxes,ha='center', va='center', fontsize=5)
        ax.text((xloc_scale+margen_total)-(tamano1)*2 , 0.03+yloc_scale, '1.0', transform=ax.transAxes,ha='center', va='center', fontsize=5)

    rect = Rectangle(((xloc_scale+margen_total)-(tamano1)*2, yloc_scale), tamano1, 0.01, fill=True, edgecolor='black', facecolor='black', linewidth=0.8, transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    
    # Tamano de 0.0km a 0.5km (son 0.5km para colorear en negro)
    if Separa_x == 1:
        ax.text((xloc_scale+margen_total)-(tamano1)*3 , 0.03+yloc_scale, '.25', transform=ax.transAxes,ha='center', va='center', fontsize=5)
        ax.text((xloc_scale+margen_total)-(tamano1)*4 , 0.03+yloc_scale, '0', transform=ax.transAxes,ha='center', va='center', fontsize=5)

    if Separa_x == 2:
        ax.text((xloc_scale+margen_total)-(tamano1)*3 , 0.03+yloc_scale, '0.5', transform=ax.transAxes,ha='center', va='center', fontsize=5)
        ax.text((xloc_scale+margen_total)-(tamano1)*4 , 0.03+yloc_scale, '0', transform=ax.transAxes,ha='center', va='center', fontsize=5)

    rect = Rectangle(((xloc_scale+margen_total)-(tamano1)*4, yloc_scale), tamano1, 0.01, fill=True, edgecolor='black', facecolor='black', linewidth=0.8, transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    # Tamano de 0.25km a 0.5km (son 025km para colorear en negro)
    tamano2 = (len_marginx + len_marginx/4)/16 
    
    if Separa_x == 1:
        ax.text(xloc_scale, 0.03+yloc_scale, '.25', transform=ax.transAxes,ha='center', va='center', fontsize=5)
        
    if Separa_x == 2:
        ax.text(xloc_scale, 0.03+yloc_scale, '0.5', transform=ax.transAxes,ha='center', va='center', fontsize=5)
    
    rect = Rectangle((xloc_scale+tamano2*2, yloc_scale), tamano2, 0.01, fill=True, edgecolor='black', facecolor='black', linewidth=0.8, transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    rect = Rectangle((xloc_scale, yloc_scale), tamano2, 0.01, fill=True, edgecolor='black', facecolor='black', linewidth=0.8, transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    
    ax.tick_params(axis='both', which='major', labelsize=8)
    plt.subplots_adjust(wspace=0.0)
    
    return fig   

#%% generar mapa secciones
def mapa_gen_scc(Cientific,Separa_x,Separa_y,COD_mun,CP_Name,seccion_shp,area_shp,map_data,map_data_limits,variable,viridis_inverted,Cmap_text,title1,title2,Title_FontSize,Cbar_Scale,Cbar_LabelSize,Cbar_FontSize,User_min_lon,User_max_lon,User_min_lat,User_max_lat):
        
    fig = plt.figure(figsize=(10.6, 6))
    fig.set_facecolor('white')
    
    # Los valores son [left, bottom, width, height] en fracciones de la figura
    rect_ax = [0.05, 0.1, 0.65, 0.8]
    ax = fig.add_axes(rect_ax)
    
    # 1). Definir el tamaño de las margenes
    # .... Cargar geopandas para el limite del mapa ...........................
    datadf = pd.read_csv(os.path.join(os.getcwd(),'Map_Limits.csv')) 
    
    # .... Aplicar la función corregida a cada fila de la columna 'geometry' 
    # y crear una lista de todas las coordenadas ..............................
    all_coords_safe = []
    datadf['geometry'].apply(lambda x: all_coords_safe.extend(extract_coordinates_safe(x)))
    # .... Calcular los valores máximos y mínimos de longitud y latitud .......
    max_lat = max(all_coords_safe, key=lambda x: x[1])[1]
    min_lat = min(all_coords_safe, key=lambda x: x[1])[1]
    max_lon = max(all_coords_safe, key=lambda x: x[0])[0]
    min_lon = min(all_coords_safe, key=lambda x: x[0])[0]
    
    # ======================= LINEA MODIFICA EL USUARIO =======================
    max_lon_new = np.ceil((min_lon) * 50) / 50 + User_min_lon
    min_lon_new = np.floor((max_lon) * 50) / 50 + User_max_lon
    
    max_lat_new = np.floor((min_lat) * 50) / 50 + User_min_lat
    min_lat_new = np.ceil((max_lat) * 50) / 50 + User_max_lat
    # =========================================================================
    
    if Separa_x == 1:
        separa_num_x = 0.01
    
    if Separa_x == 2:
        separa_num_x = 0.02
        
    if Separa_y == 1:
        separa_num_y = 0.01
    
    if Separa_y == 2:
        separa_num_y = 0.02
    
    ticks_x = np.arange(min_lon_new, max_lon_new - separa_num_x, - separa_num_x)
    ticks_y = np.arange(min_lat_new, max_lat_new - separa_num_y, - separa_num_y)
    
    # Dependiendo de la separación de longitud o latitud
    len_marginx = LenMargin_value(ticks_x)*Separa_x    # Longitud margen en x
    len_marginy = LenMargin_value(ticks_y)*Separa_y    # Longitud margen en y

    
    
    # 2). Definir posicion de ax_legend
    #Ecuacion_linea =-0.2037647997871495 x + 0.825043235333245
    
    ax_left = np.around(-0.2037647997871495*(len_marginx/len_marginy) + 0.825043235333245,2)+0.01
    
    rect_ax_legend = [ax_left, 0.1, 0.18, 0.8]
    ax_legend = fig.add_axes(rect_ax_legend)
    
    
    # ax -- axis del mapa
    # ax_legend -- axis para presentacion del mapa
    # =========================================================================
    
    # .... Cargar los shape files .............................................
    valor_maximo = map_data[variable].max()
    
    if valor_maximo < 1:
        norm = Normalize(vmin=0, vmax=1)
        map_data.plot(column=variable, ax=ax, edgecolor='grey', alpha=1.0, cmap=viridis_inverted, linewidth=0.4, norm=norm)
    else:
        map_data.plot(column=variable, ax=ax, edgecolor='grey', alpha=1.0, cmap=viridis_inverted, linewidth=0.4)
        
    seccion_shp.plot(ax=ax, edgecolor='black', facecolor="none", alpha=1.0, linewidth=0.5)
    area_shp.plot(ax=ax, edgecolor='black', facecolor="none", alpha=1.0, linewidth=1.0)
    # .... Brujula señalando el norte .........................................
    x, y, arrow_length = 0.95, 0.95, 0.1
    ax.annotate('N', xy=(x, y), xytext=(x, y-arrow_length),
                arrowprops=dict(facecolor='black', width=3, headwidth=9),
                ha='center', va='center', fontsize=11, xycoords=ax.transAxes)
    # .... Cbar - Resultados Riesgo ...........................................
    # Calcula la varianza de los datos para verificar si todos los valores son iguales
    varianza_datos = np.var(map_data[variable])
    if varianza_datos == 0:
        # Todos los valores son iguales
        
        # lista de valores únicos, incluidos los NaN
        valores_unicos = map_data[variable].unique()
        # Filtra los valores, eliminando los NaN
        valores_sin_nan = valores_unicos[~np.isnan(valores_unicos)]
        if valores_sin_nan.size == 1:
            valor_unico = valores_sin_nan[0]
        else:
            # Manejar el caso de múltiples valores o ningún valor después de eliminar NaN
            valor_unico = None  
    
        norm = Normalize(vmin=valor_unico, vmax=valor_unico)
    
        # Ajusta las dimensiones del cax si es necesario
        cax = fig.add_axes([rect_ax_legend[0]+0.05, 0.270, 0.08, 0.015]) # [left, bottom, width, height]
        
        # Obtener el colormap viridis
        viridis = plt.cm.get_cmap('viridis', 256)
        
        # El primer color de viridis invertido es el último color de viridis
        primer_color_viridis_inverted = viridis.colors[-1]
        
        # Crear un colormap personalizado que consiste solo en ese color
        colormap_personalizado = ListedColormap([primer_color_viridis_inverted])
        
        # Crea el colorbar con un solo color
        cbar = fig.colorbar(cm.ScalarMappable(norm=norm, cmap=colormap_personalizado), cax=cax, orientation='horizontal')
        
        # Establece un solo tick con el valor único
        cbar.set_ticks([valor_unico])
        cbar.ax.tick_params(axis='x', labelsize=Cbar_LabelSize)
        
    else:
        if valor_maximo < 1:
            cax = fig.add_axes([rect_ax_legend[0]+0.03, 0.270, 0.115, 0.015]) # Ajusta como sea necesario
            cbar = fig.colorbar(cm.ScalarMappable(norm=norm, cmap=viridis_inverted), cax=cax, orientation='horizontal')
            valores = np.linspace(0, 1, 5)
            
        else:
            norm = Normalize(vmin=np.floor(map_data[variable].min()), vmax=np.ceil(map_data[variable].max()))
        
            cax = fig.add_axes([rect_ax_legend[0]+0.03, 0.270, 0.115, 0.015]) # Ajusta como sea necesario
            cbar = fig.colorbar(cm.ScalarMappable(norm=norm, cmap=viridis_inverted), cax=cax, orientation='horizontal')
        
            valores = np.linspace(np.floor(map_data[variable].min()), np.ceil(map_data[variable].max()), 6)
            
        if Cientific == 1:
            formatter = FuncFormatter(scientific_formatter)
            cbar.set_ticks(np.round(valores[1:-1] / Cbar_Scale) * Cbar_Scale)
            cbar.ax.xaxis.set_major_formatter(formatter)
            cbar.ax.tick_params(axis='x', labelsize=Cbar_LabelSize)
        else:
            cbar.set_ticks(np.round(valores[1:-1] / Cbar_Scale) * Cbar_Scale)
            cbar.ax.tick_params(axis='x', labelsize=Cbar_LabelSize)
    
    ax_legend.text(0.5, 0.25, Cmap_text, transform=ax_legend.transAxes,ha='center', va='center', fontsize=Cbar_FontSize)
    # .... Eliminar los ejes de ax_legend .....................................
    ax_legend.set_axis_off()
    # .... En ax_legend anadir los logos de ACOFI y del servicio .............. 
    # Logo ACOFI
    image_path = os.path.join(os.path.join(os.getcwd(), "icon"),'acofilogo.png')
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.0215)
    ab = offsetbox.AnnotationBbox(imagebox, (0.25, 0.95), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)
    # Logo Servicio Geologico
    image_path = os.path.join(os.path.join(os.getcwd(), "icon"),'serviciogc.png')
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.0515)
    ab = offsetbox.AnnotationBbox(imagebox, (0.75, 0.95), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)
    
    
    # .... Generar recuadros para la leyenda ..................................
    ax_legend.add_patch(Rectangle((0.02, 0.785), 0.96, 0.10, fill=True, edgecolor='grey', facecolor='white', linewidth=1.0, transform=ax_legend.transAxes, clip_on=False))
    ax_legend.text(0.5, 0.85, 'Modelo Nacional', transform=ax_legend.transAxes,ha='center', va='center', fontsize=7.8,fontweight='bold')
    ax_legend.text(0.5, 0.82, 'de Riesgo Sísmico', transform=ax_legend.transAxes,ha='center', va='center', fontsize=7.8,fontweight='bold')
    
    ax_legend.add_patch(Rectangle((0.02, 0.67), 0.96, 0.10, fill=True, edgecolor='grey', facecolor='white', linewidth=1.0, transform=ax_legend.transAxes, clip_on=False))
    ax_legend.text(0.5, 0.75, COD_mun, transform=ax_legend.transAxes,ha='center', va='center', fontsize=6,fontweight='bold')
    ax_legend.text(0.5, 0.71, CP_Name, transform=ax_legend.transAxes,ha='center', va='center', fontsize=9,fontweight='bold')
    
    ax_legend.add_patch(Rectangle((0.02, 0.555), 0.96, 0.10, fill=True, edgecolor='grey', facecolor='white', linewidth=1.0, transform=ax_legend.transAxes, clip_on=False))
    ax_legend.text(0.5, 0.615, title1, transform=ax_legend.transAxes,ha='center', va='center', fontsize=Title_FontSize,fontweight='bold')
    ax_legend.text(0.5, 0.59, title2, transform=ax_legend.transAxes,ha='center', va='center', fontsize=Title_FontSize,fontweight='bold')
    
    ax_legend.add_patch(Rectangle((0.02, 0.1), 0.96, 0.44, fill=True, edgecolor='grey', facecolor='white', linewidth=1.0, transform=ax_legend.transAxes, clip_on=False))
    ax_legend.text(0.235, 0.5, 'Leyenda', transform=ax_legend.transAxes,ha='center', va='center', fontsize=8.5)
    ax_legend.text(0.26, 0.45, 'MGN (DANE)', transform=ax_legend.transAxes,ha='center', va='center', fontsize=6,fontweight='bold')
    
    # ax_legend.add_patch(Rectangle((0.715, 0.43), 0.025, 0.015, edgecolor='black', facecolor='none', transform=fig.transFigure))

    
    ax_legend.add_patch(Rectangle((0.105, 0.405), 0.125, 0.02, fill=True, edgecolor='black', facecolor='white', linewidth=1.4, transform=ax_legend.transAxes, clip_on=False))
    ax_legend.text(0.40, 0.412, 'Área censal', transform=ax_legend.transAxes,ha='center', va='center', fontsize=6)
    ax_legend.add_patch(Rectangle((0.105, 0.368), 0.125, 0.02, fill=True, edgecolor='black', facecolor='white', linewidth=0.9, transform=ax_legend.transAxes, clip_on=False))
    ax_legend.text(0.44, 0.375, 'Sección urbana', transform=ax_legend.transAxes,ha='center', va='center', fontsize=6)
    
    ax_legend.text(0.34, 0.29, 'Resultados riesgo', transform=ax_legend.transAxes,ha='center', va='center', fontsize=6,fontweight='bold')
    
    image_path = os.path.join(os.path.join(os.getcwd(), "icon"),'logosabana.png')
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.17)
    ab = offsetbox.AnnotationBbox(imagebox, (0.25, 0.05), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)
    
    image_path = os.path.join(os.path.join(os.getcwd(), "icon"),'logomedellin.png')
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.12)
    ab = offsetbox.AnnotationBbox(imagebox, (0.75, 0.05), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)
    
    # .... Añadir mapa base ...................................................
    ctx.add_basemap(ax, crs=map_data.crs.to_string(), source=ctx.providers.CartoDB.Positron, zoom=13, alpha=0.5)
    
    
    ax.set_xlim([np.ceil((min_lon) * 50) / 50 + User_min_lon ,np.floor((max_lon) * 50) / 50 + User_max_lon])
    ax.set_ylim([np.floor((min_lat) * 50) / 50 + User_min_lat , np.ceil((max_lat) * 50) / 50 + User_max_lat])
    
    ax.set_xticks(ticks_x)
    ax.set_yticks(ticks_y)
    

    # ============ LINEA PARA MODIFICAR SEGUN CADA CASO =======================
    # .... Generar margenes del mapa ..........................................
    suma = 0
    for index in range(len(ticks_x)-1):
        if suma%2 == 0:
            rect = Rectangle((len_marginx*suma, 0.0), len_marginx, 0.008, fill=True, edgecolor='black', facecolor='black', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((len_marginx*suma, 1.0-0.008), len_marginx, 0.008, fill=True, edgecolor='black', facecolor='black', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1
        else:
            rect = Rectangle((len_marginx*suma, 0.0), len_marginx, 0.008, fill=True, edgecolor='black', facecolor='white', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((len_marginx*suma, 1.0-0.008), len_marginx, 0.008, fill=True, edgecolor='black', facecolor='white', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1
    
    suma = 0
    for index in range(len(ticks_y)-1):
        if suma%2 == 0:
            rect = Rectangle((0.0, len_marginy*suma), 0.007, len_marginy, fill=True, edgecolor='black', facecolor='black', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((1.0 - 0.007, len_marginy*suma), 0.007, len_marginy, fill=True, edgecolor='black', facecolor='black', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1
        else:
            rect = Rectangle((0.0, len_marginy*suma), 0.007, len_marginy, fill=True, edgecolor='black', facecolor='white', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((1.0 - 0.007, len_marginy*suma), 0.007, len_marginy, fill=True, edgecolor='black', facecolor='white', linewidth=0.5, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1
    # =========================================================================
    
    xloc_scale = 0.58
    yloc_scale = 0.08
    
    # =========================================================================
    
    # .... Generar grilla .....................................................
    ax.grid(True, which='both', color='grey', linewidth=0.8, linestyle='-', alpha=0.2)
    len_marginx = len_marginx-len_marginx/5
    # .... Tamano total de la barra escala ....................................
    # Tamano total de la escala (2km + 0.5km)
    # (len_marginx + len_marginx/4) son 2km --> 0.5km son (len_marginx + len_marginx/4)/4
    margen_total = (len_marginx + len_marginx/4) + (len_marginx + len_marginx/4)/4
    
    if Separa_x == 1:
        ax.text((xloc_scale+0.01)+margen_total, 0.03+yloc_scale, '1km', transform=ax.transAxes,ha='center', va='center', fontsize=6)
        
    if Separa_x == 2:
        ax.text((xloc_scale+0.01)+margen_total, 0.03+yloc_scale, '2km', transform=ax.transAxes,ha='center', va='center', fontsize=6)
    
    rect = Rectangle((xloc_scale, yloc_scale), margen_total, 0.011, fill=True, edgecolor='black', facecolor='white', linewidth=0.8, transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    # Tamano de 1km a 1.5km (son 0.5km para colorear en negro)
    tamano1 = (len_marginx + len_marginx/4)/4
    
    if Separa_x == 1:
        ax.text((xloc_scale+margen_total)-(tamano1) , 0.03+yloc_scale, '.75', transform=ax.transAxes,ha='center', va='center', fontsize=5)
        ax.text((xloc_scale+margen_total)-(tamano1)*2 , 0.03+yloc_scale, '0.5', transform=ax.transAxes,ha='center', va='center', fontsize=5)

    if Separa_x == 2:
        ax.text((xloc_scale+margen_total)-(tamano1) , 0.03+yloc_scale, '1.5', transform=ax.transAxes,ha='center', va='center', fontsize=5)
        ax.text((xloc_scale+margen_total)-(tamano1)*2 , 0.03+yloc_scale, '1.0', transform=ax.transAxes,ha='center', va='center', fontsize=5)

    rect = Rectangle(((xloc_scale+margen_total)-(tamano1)*2, yloc_scale), tamano1, 0.01, fill=True, edgecolor='black', facecolor='black', linewidth=0.8, transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    
    # Tamano de 0.0km a 0.5km (son 0.5km para colorear en negro)
    if Separa_x == 1:
        ax.text((xloc_scale+margen_total)-(tamano1)*3 , 0.03+yloc_scale, '.25', transform=ax.transAxes,ha='center', va='center', fontsize=5)
        ax.text((xloc_scale+margen_total)-(tamano1)*4 , 0.03+yloc_scale, '0', transform=ax.transAxes,ha='center', va='center', fontsize=5)

    if Separa_x == 2:
        ax.text((xloc_scale+margen_total)-(tamano1)*3 , 0.03+yloc_scale, '0.5', transform=ax.transAxes,ha='center', va='center', fontsize=5)
        ax.text((xloc_scale+margen_total)-(tamano1)*4 , 0.03+yloc_scale, '0', transform=ax.transAxes,ha='center', va='center', fontsize=5)

    rect = Rectangle(((xloc_scale+margen_total)-(tamano1)*4, yloc_scale), tamano1, 0.01, fill=True, edgecolor='black', facecolor='black', linewidth=0.8, transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    # Tamano de 0.25km a 0.5km (son 025km para colorear en negro)
    tamano2 = (len_marginx + len_marginx/4)/16 
    
    if Separa_x == 1:
        ax.text(xloc_scale, 0.03+yloc_scale, '.25', transform=ax.transAxes,ha='center', va='center', fontsize=5)
        
    if Separa_x == 2:
        ax.text(xloc_scale, 0.03+yloc_scale, '0.5', transform=ax.transAxes,ha='center', va='center', fontsize=5)
    
    rect = Rectangle((xloc_scale+tamano2*2, yloc_scale), tamano2, 0.01, fill=True, edgecolor='black', facecolor='black', linewidth=0.8, transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    rect = Rectangle((xloc_scale, yloc_scale), tamano2, 0.01, fill=True, edgecolor='black', facecolor='black', linewidth=0.8, transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    
    ax.tick_params(axis='both', which='major', labelsize=8)
    plt.subplots_adjust(wspace=0.0)
    
    return fig   
#%% ====== Funcion Danos ======================================================
def Gen_Tabla_Resume_DNO(Num_build,aggrisk):
    
    wb = Workbook()
    ws = wb.active
    # Configuracion de la primera fila
    ws.merge_cells('B2:C2')
    ws['B2'] = 'Número total de edificaciones'
    ws['D2'] = Num_build
    # Configuración de la segunda fila
    ws.merge_cells('B3:B4')
    ws['B3'] = 'Estado de daño'
    ws.merge_cells('C3:D3')
    ws['C3'] = 'Número de edificaciones'
    # Configuración de la terecera fila
    ws['C4'] = '[-]'
    ws['D4'] = '[%]'
    # Configuracion estados de daño
    title_dmg = ['Sin daño','Daño leve','Daño moderado','Daño severo','Daño completo/colapso']
    sumaprc = []
    suma = []
    for index, dmg in enumerate(aggrisk):
        ws['B'+str(index+5)] = title_dmg[index]
        ws['C'+str(index+5)] = np.around(dmg,0)
        ws['D'+str(index+5)] = np.around(dmg/np.sum(aggrisk)*100,2)
        suma.append(np.around(dmg,0))
        sumaprc.append(np.around(dmg/np.sum(aggrisk)*100,2))
    ws['B10'] = 'Total'
    ws['C10'] = np.sum(suma)
    ws['D10'] = str(np.around(np.sum(sumaprc),0))+'%'
    # Alinear el contenido de las celdas al centro
    for row in ws.iter_rows(min_row=2, max_row=10, min_col=2, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    # Configurar el tamaño de las celdas
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 11

    # Agregar bordes a todas las celdas
    for row in ws.iter_rows(min_row=2, max_row=10, min_col=2, max_col=4):
        for cell in row:
            cell.border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
    # Guardar la tabla de resumen
    Table_Resu = wb.save
    
    return Table_Resu

def Gen_Tabla_taxonomia_DNO(df_expotax):
    wb = Workbook()
    ws = wb.active
    # Configuracion de la primera fila
    ws.merge_cells('B2:B4')
    ws['B2'] = 'Tipología constructiva'
    ws.merge_cells('C2:L2')
    ws['C2'] = 'Número de edificaciones'
    # Configuración de la segunda fila
    ws.merge_cells('C3:D3')
    ws['C3'] = 'Sin daño'
    ws.merge_cells('E3:F3')
    ws['E3'] = 'Daño leve'
    ws.merge_cells('G3:H3')
    ws['G3'] = 'Daño moderado'
    ws.merge_cells('I3:J3')
    ws['I3'] = 'Daño severo'
    ws.merge_cells('K3:L3')
    ws['K3'] = 'Daño completo'
    # Configuración de la tercera fila 
    columns1 = ['C','E','G','I','K']
    columns2 = ['D','F','H','J','L']
    for col in columns1:
        ws[col+'4'] = '[-]'
    for col in columns2:
        ws[col+'4'] = '[%]'
    # Configuracion taxonomias
    title_dmg = ['Sin daño','Daño leve','Daño moderado','Daño severo','Daño completo/colapso']
    suma0, suma1, suma2, suma3, suma4 = [], [], [], [], []
    for index in range(len(df_expotax)):
        ws['B'+str(index+5)] = df_expotax.taxonomy[index]
        ws['C'+str(index+5)] = np.around(df_expotax.dmg0[index],0)
        ws['D'+str(index+5)] = np.around(df_expotax.dmg0[index]/np.sum(df_expotax.values[index][1:-1])*100,2)
        ws['E'+str(index+5)] = np.around(df_expotax.dmg1[index],0)
        ws['F'+str(index+5)] = np.around(df_expotax.dmg1[index]/np.sum(df_expotax.values[index][1:-1])*100,2)
        ws['G'+str(index+5)] = np.around(df_expotax.dmg2[index],0)
        ws['H'+str(index+5)] = np.around(df_expotax.dmg2[index]/np.sum(df_expotax.values[index][1:-1])*100,2)
        ws['I'+str(index+5)] = np.around(df_expotax.dmg3[index],0)
        ws['J'+str(index+5)] = np.around(df_expotax.dmg3[index]/np.sum(df_expotax.values[index][1:-1])*100,2)
        ws['K'+str(index+5)] = np.around(df_expotax.dmg4[index],0)
        ws['L'+str(index+5)] = np.around(df_expotax.dmg4[index]/np.sum(df_expotax.values[index][1:-1])*100,2)
        suma0.append(np.around(df_expotax.dmg0[index],2))
        suma1.append(np.around(df_expotax.dmg1[index],2))
        suma2.append(np.around(df_expotax.dmg2[index],2))
        suma3.append(np.around(df_expotax.dmg3[index],2))
        suma4.append(np.around(df_expotax.dmg4[index],2))
        
    # ws.merge_cells('B'+str(index+6)+':C'+str(index+6))
    ws['B'+str(index+6)] = 'Total'
    total_edif = np.around(np.sum(suma0)+np.sum(suma1)+np.sum(suma2)+np.sum(suma3)+np.sum(suma4),0)
    ws['C'+str(index+6)] = np.around(np.sum(suma0),0)
    ws['D'+str(index+6)] = str(np.around(np.sum(suma0)/total_edif*100,2))+'%'

    ws['E'+str(index+6)] = np.around(np.sum(suma1),0)
    ws['F'+str(index+6)] = str(np.around(np.sum(suma1)/total_edif*100,2))+'%'

    ws['G'+str(index+6)] = np.around(np.sum(suma2),0)
    ws['H'+str(index+6)] = str(np.around(np.sum(suma2)/total_edif*100,2))+'%'

    ws['I'+str(index+6)] = np.around(np.sum(suma3),0)
    ws['J'+str(index+6)] = str(np.around(np.sum(suma3)/total_edif*100,2))+'%'

    ws['K'+str(index+6)] = np.around(np.sum(suma4),0)
    ws['L'+str(index+6)] = str(np.around(np.sum(suma4)/total_edif*100,2))+'%'


    # Alinear el contenido de las celdas al centro
    for row in ws.iter_rows(min_row=2, max_row=len(df_expotax)+5, min_col=2, max_col=12):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    # Configurar el tamaño de las celdas
    ws.column_dimensions['B'].width = 12.22
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 6.7
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 6.33
    ws.column_dimensions['G'].width = 5
    ws.column_dimensions['H'].width = 6.33
    ws.column_dimensions['I'].width = 5
    ws.column_dimensions['J'].width = 6.33
    ws.column_dimensions['K'].width = 5
    ws.column_dimensions['L'].width = 6.33

    # Agregar bordes a todas las celdas
    for row in ws.iter_rows(min_row=2, max_row=len(df_expotax)+5, min_col=2, max_col=12):
        for cell in row:
            cell.border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
            
    # Guardar el archivo 
    Table_Resu = wb.save
    
    return Table_Resu

def Figura_taxonomias_DNO(taxonomias,aggrisk_mnz,canvas_master,relex,reley):
    categorias = range(1,len(taxonomias)+1)
    sin_danio,danio_leve,danio_moderado,danio_extensivo,colapso = [], [], [], [], []
    for index in range(len(taxonomias)):
        sindano = (aggrisk_mnz[index,0]/np.sum(aggrisk_mnz[index,:]))*100
        sin_danio.append(sindano)
        leve = (aggrisk_mnz[index,1]/np.sum(aggrisk_mnz[index,:]))*100
        danio_leve.append(leve)
        moderado = (aggrisk_mnz[index,2]/np.sum(aggrisk_mnz[index,:]))*100
        danio_moderado.append(moderado)
        severo = (aggrisk_mnz[index,3]/np.sum(aggrisk_mnz[index,:]))*100
        danio_extensivo.append(severo)
        colaps = (aggrisk_mnz[index,4]/np.sum(aggrisk_mnz[index,:]))*100
        colapso.append(colaps)

    bar_width = 0.75
    indices = np.arange(len(categorias))
    yticks = []
    for index in np.arange(0, 101, 10):
        yticks.append(str(index)+'%')
        
    
    fig, ax = plt.subplots(figsize=(7,4.4))
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.95, bottom=0.19)
    
    # Configuración de las barras apiladas
    p5 = ax.bar(indices, colapso, bar_width, bottom=np.array(sin_danio) + np.array(danio_leve) + np.array(danio_moderado) + np.array(danio_extensivo), color='#FF3B3B', label='Colapso')
    p4 = ax.bar(indices, danio_extensivo, bar_width, bottom=np.array(sin_danio) + np.array(danio_leve) + np.array(danio_moderado), color='#F2A068', label='Daño extensivo')
    p3 = ax.bar(indices, danio_moderado, bar_width, bottom=np.array(sin_danio) + np.array(danio_leve), color='#FFD966', label='Daño moderado')
    p2 = ax.bar(indices, danio_leve, bar_width, bottom=sin_danio, color='#92D050', label='Daño leve')
    p1 = ax.bar(indices, sin_danio, bar_width, color='#D0CECE', label='Sin daño', alpha=0.6)

    # Etiquetas y leyendas
    ax.set_xlabel('Taxonomías',fontsize=4)
    ax.set_ylabel('Edificaciones en estado de daño [%]',fontsize=4)
    ax.set_xticks(indices, categorias)

    ax.set_yticks(np.arange(0, 101, 10))
    ax.set_yticklabels(yticks,fontsize=3.5)
    ax.set_xticklabels(categorias,rotation=0,fontsize=3.5)
    # fig.legend(loc='lower center', bbox_to_anchor=(0.5, 0.013), ncol=3, fontsize=8)
    
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(fig, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False)
    canvas.get_tk_widget().place(relx=relex, rely=reley, anchor=tk.CENTER, width=800, height=500)
    
    return canvas

def Figura_taxonomias_DNO_Export(taxonomias,aggrisk_mnz):
    categorias = taxonomias
    sin_danio,danio_leve,danio_moderado,danio_extensivo,colapso = [], [], [], [], []
    for index in range(len(taxonomias)):
        sindano = (aggrisk_mnz[index,0]/np.sum(aggrisk_mnz[index,:]))*100
        sin_danio.append(sindano)
        leve = (aggrisk_mnz[index,1]/np.sum(aggrisk_mnz[index,:]))*100
        danio_leve.append(leve)
        moderado = (aggrisk_mnz[index,2]/np.sum(aggrisk_mnz[index,:]))*100
        danio_moderado.append(moderado)
        severo = (aggrisk_mnz[index,3]/np.sum(aggrisk_mnz[index,:]))*100
        danio_extensivo.append(severo)
        colaps = (aggrisk_mnz[index,4]/np.sum(aggrisk_mnz[index,:]))*100
        colapso.append(colaps)

    bar_width = 0.75
    indices = np.arange(len(categorias))
    yticks = []
    for index in np.arange(0, 101, 10):
        yticks.append(str(index)+'%')
        
    
    fig, ax = plt.subplots(figsize=(7,4.4))
    plt.subplots_adjust(left=0.15, right=0.85, top=0.99, bottom=0.35)
    
    # Configuración de las barras apiladas
    p5 = ax.bar(indices, colapso, bar_width, bottom=np.array(sin_danio) + np.array(danio_leve) + np.array(danio_moderado) + np.array(danio_extensivo), color='#FF3B3B', label='Colapso')
    p4 = ax.bar(indices, danio_extensivo, bar_width, bottom=np.array(sin_danio) + np.array(danio_leve) + np.array(danio_moderado), color='#F2A068', label='Daño extensivo')
    p3 = ax.bar(indices, danio_moderado, bar_width, bottom=np.array(sin_danio) + np.array(danio_leve), color='#FFD966', label='Daño moderado')
    p2 = ax.bar(indices, danio_leve, bar_width, bottom=sin_danio, color='#92D050', label='Daño leve')
    p1 = ax.bar(indices, sin_danio, bar_width, color='#D0CECE', label='Sin daño', alpha=0.6)

    # Etiquetas y leyendas
    ax.set_xlabel('Taxonomías',fontsize=8)
    ax.set_ylabel('Porcentaje de edificaciones en un estado de daño',fontsize=7)
    ax.set_xticks(indices, categorias)

    ax.set_yticks(np.arange(0, 101, 10))
    ax.set_yticklabels(yticks,fontsize=8)
    ax.set_xticklabels(categorias,rotation=90,fontsize=6.5)
    fig.legend(loc='lower center', bbox_to_anchor=(0.5, 0.013), ncol=3, fontsize=8)

    return fig


#%% ====== Funciones MAPAS Danos ==============================================

def canva_DNO_ocupantes_mnz(COD_mun,CP_Name,manzana_shp,seccion_shp,area_shp,map_data,Modelo_Expo,canvas_master,relex,reley,Separa_x,Separa_y,User_min_lon,User_max_lon,User_min_lat,User_max_lat,variable):
    
    # =================== VARIABLES MODIFICAR - area_construida ===================
    Title_FontSize = 6.3         # Tamano de letra del nombre de la variable presentada en el mapa
    Cbar_LabelSize = 4           # Tamano de letra de los ticks del cbar
    Cbar_FontSize = 6.5          # Tamano de letra del titulo del cbar
    Cientific = 0                # Notacion cientifica 1: si 0: no
    # =============================================================================
    
    viridis_inverted = ["a","b"]
    Cbar_Scale = ["a","b"]
    Cmap_text = ["a","b"]
    title1 = ["a","b"]
    title2 = ["a","b"]
    
    if variable.startswith("aac") and variable.endswith("hab"):
        Cbar_Scale[1] = 0.003           # Espaciado de los ticks del cbar
        Cmap_text[1] = 'Ocupantes en edificios colapsados'
        title1[1] = 'Número anual promedio de ocupantes en'
        title2[1] = 'edificios colapsados por manzana censal'
        viridis_inverted[1] = 'GnBu'
    elif variable.startswith("aac") and variable.endswith("edis"):
        Cbar_Scale[1] = 0.003           # Espaciado de los ticks del cbar
        Cmap_text[1] = 'Edificios colapsados'
        title1[1] = 'Número anual promedio de edificios'
        title2[1] = 'colapsados por manzana censal'
        viridis_inverted[1] = 'GnBu'
    elif variable.startswith("aad") and variable.endswith("heridos_hab"):
        Cbar_Scale[1] = 0.003           # Espaciado de los ticks del cbar
        Cmap_text[1] = 'Heridos anuales promedio'
        title1[1] = 'Heridos anuales promedio'
        title2[1] = 'para cada manzana censal'
        viridis_inverted[1] = 'PuRd'
    elif variable.startswith("aad") and variable.endswith("fallecidos_hab"):
        Cbar_Scale[1] = 0.002           # Espaciado de los ticks del cbar
        Cmap_text[1] = 'Fallecidos anuales promedio'
        title1[1] = 'Fallecidos anuales promedio'
        title2[1] = 'por cada manzana censal'
        viridis_inverted[1] = 'Reds'

    # Generar dataframe data_figura38
    # Para eso necesito:
        # 1. Modelo de exposicion del municipio (informar en guia el nombre de las columnas)
        
    # Codigo de las manzanas corregido

    cod_mnzdef_model = []
    for mnz in Modelo_Expo.cod_mnz:
        if mnz[0].isalpha():
            mnz2 = str(mnz)
            cod_mnzdef_model.append(mnz2[1::])   
        else:
            mnz2 = str(mnz)
            cod_mnzdef_model.append(mnz2)  
            
    Modelo_Expo.cod_mnz = cod_mnzdef_model

    cod_sccdef_model = []
    for scc in Modelo_Expo.cod_secc:
        if scc[0].isalpha():
            cod_sccdef_model.append(scc[1::])
        else:
            cod_sccdef_model.append(scc)
               
    Modelo_Expo.cod_secc = cod_sccdef_model
    
    # Obtener taxonomia representativa
    Expo_groupby_txn = Modelo_Expo.groupby('tipologia')['area_cons'] 
    Expo_groupby_mnz = Modelo_Expo.groupby('cod_mnz')['area_cons'].sum()

    Area_constr_txn = Expo_groupby_txn.sum()
    df_area = pd.DataFrame({'tipologia':Area_constr_txn.index,'area':Area_constr_txn})
    Txn_Rep = df_area.loc[df_area.area == df_area.area.max()].values[0][0]          
    # Obtener el % de area construida de las manzanas con esa taxonomia representativa
    df_mnz_txn = Modelo_Expo.loc[Modelo_Expo.tipologia == Txn_Rep]
    df_mnz_groupbymnz = df_mnz_txn.groupby('cod_mnz')['area_cons'].sum()
    df_final = pd.DataFrame({'cod_mnz2':df_mnz_groupbymnz.index,'area':df_mnz_groupbymnz})
    df_final = df_final.merge(Expo_groupby_mnz, left_on='cod_mnz2', right_on='cod_mnz', how='left')
    df_final['area_cons'] = (df_final.area/df_final.area_cons)*100 


    # Geopandas para definir los limites del mapa
    map_data_limits = df_final.merge(manzana_shp, left_on='cod_mnz2', right_on='COD_DANE', how='left')
    map_data_limits_df = map_data_limits.to_csv("Map_Limits.csv",index=False)

    
    # Generar mapa de manzanas
    figura = mapa_gen_mnz(Cientific,Separa_x,Separa_y,COD_mun,CP_Name,manzana_shp,seccion_shp,area_shp,map_data,map_data_limits_df,variable,viridis_inverted[1],Cmap_text[1],title1[1],title2[1],Title_FontSize,Cbar_Scale[1],Cbar_LabelSize,Cbar_FontSize,User_min_lon,User_max_lon,User_min_lat,User_max_lat)
    
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(figura, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False)
    canvas.get_tk_widget().place(relx=relex, rely=reley, anchor=tk.CENTER, width=972, height=550)
    
    return canvas

def canva_DNO_ocupantes_scc(COD_mun,CP_Name,seccion_shp,area_shp,map_data,Modelo_Expo,canvas_master,relex,reley,Separa_x,Separa_y,User_min_lon,User_max_lon,User_min_lat,User_max_lat,variable):
    
    # =================== VARIABLES MODIFICAR - area_construida ===================
    Title_FontSize = 6.5         # Tamano de letra del nombre de la variable presentada en el mapa
    Cbar_LabelSize = 4           # Tamano de letra de los ticks del cbar
    Cbar_FontSize = 6.5          # Tamano de letra del titulo del cbar
    Cientific = 0                # Notacion cientifica 1: si 0: no
    # =============================================================================
    
    viridis_inverted = ["a","b"]
    Cbar_Scale = ["a","b"]
    Cmap_text = ["a","b"]
    title1 = ["a","b"]
    title2 = ["a","b"]
    
    if variable.startswith("aac") and variable.endswith("hab"):
        Cbar_Scale[0] = 0.003           # Espaciado de los ticks del cbar
        Cmap_text[0] = 'Ocupantes en edificios colapsados'
        title1[0] = 'Número anual promedio de ocupantes en'
        title2[0] = 'edificios colapsados por sección urbana'
        viridis_inverted[0] = 'GnBu'
    elif variable.startswith("aac") and variable.endswith("edis"):
        Cbar_Scale[0] = 0.003           # Espaciado de los ticks del cbar
        Cmap_text[0] = 'Edificios colapsados'
        title1[0] = 'Número anual promedio de edificios'
        title2[0] = 'colapsados por sección urbana'
        viridis_inverted[0] = 'GnBu'
    elif variable.startswith("aad") and variable.endswith("heridos_hab"):
        Cbar_Scale[0] = 0.003           # Espaciado de los ticks del cbar
        Cmap_text[0] = 'Heridos anuales promedio'
        title1[0] = 'Heridos anuales promedio'
        title2[0] = 'para cada sección urbana'
        viridis_inverted[0] = 'PuRd'
    elif variable.startswith("aad") and variable.endswith("fallecidos_hab"):
        Cbar_Scale[0] = 0.002           # Espaciado de los ticks del cbar
        Cmap_text[0] = 'Fallecidos anuales promedio'
        title1[0] = 'Fallecidos anuales promedio'
        title2[0] = 'por cada sección urbana'
        viridis_inverted[0] = 'Reds'

    # Generar dataframe data_figura38
    # Para eso necesito:
        # 1. Modelo de exposicion del municipio (informar en guia el nombre de las columnas)
        
    # Codigo de las manzanas corregido

    cod_mnzdef_model = []
    for mnz in Modelo_Expo.cod_mnz:
        if mnz[0].isalpha():
            mnz2 = str(mnz)
            cod_mnzdef_model.append(mnz2[1::])   
        else:
            mnz2 = str(mnz)
            cod_mnzdef_model.append(mnz2)  
            
    Modelo_Expo.cod_mnz = cod_mnzdef_model

    cod_sccdef_model = []
    for scc in Modelo_Expo.cod_secc:
        if scc[0].isalpha():
            cod_sccdef_model.append(scc[1::])
        else:
            cod_sccdef_model.append(scc)
               
    Modelo_Expo.cod_secc = cod_sccdef_model
    
    # Obtener taxonomia representativa
    Expo_groupby_txn = Modelo_Expo.groupby('tipologia')['area_cons'] 
    Expo_groupby_mnz = Modelo_Expo.groupby('cod_secc')['area_cons'].sum()

    Area_constr_txn = Expo_groupby_txn.sum()
    df_area = pd.DataFrame({'tipologia':Area_constr_txn.index,'area':Area_constr_txn})
    Txn_Rep = df_area.loc[df_area.area == df_area.area.max()].values[0][0]          
    # Obtener el % de area construida de las manzanas con esa taxonomia representativa
    df_mnz_txn = Modelo_Expo.loc[Modelo_Expo.tipologia == Txn_Rep]
    df_mnz_groupbymnz = df_mnz_txn.groupby('cod_secc')['area_cons'].sum()
    df_final = pd.DataFrame({'cod_scc2':df_mnz_groupbymnz.index,'area':df_mnz_groupbymnz})
    df_final = df_final.merge(Expo_groupby_mnz, left_on='cod_scc2', right_on='cod_secc', how='left')
    df_final['area_cons'] = (df_final.area/df_final.area_cons)*100 


    # Geopandas para definir los limites del mapa
    map_data_limits = df_final.merge(seccion_shp, left_on='cod_scc2', right_on='COD_SECC', how='left')
    map_data_limits_df = map_data_limits.to_csv("Map_Limits_secc.csv",index=False)

    
    # Generar mapa de manzanas
    figura = mapa_gen_scc(Cientific,Separa_x,Separa_y,COD_mun,CP_Name,seccion_shp,area_shp,map_data,map_data_limits_df,variable,viridis_inverted[0],Cmap_text[0],title1[0],title2[0],Title_FontSize,Cbar_Scale[0],Cbar_LabelSize,Cbar_FontSize,User_min_lon,User_max_lon,User_min_lat,User_max_lat)
    
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(figura, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False)
    canvas.get_tk_widget().place(relx=relex, rely=reley, anchor=tk.CENTER, width=972, height=550)
    
    return canvas

#%% ====== Funcion MAPAS -- Generados =========================================
def GeneradorMapas(COD_mun,CP_Name,manzana_shp,seccion_shp,area_shp,map_data,Modelo_Expo,Separa_x,Separa_y,User_min_lon,User_max_lon,User_min_lat,User_max_lat,variable,Txn_Rep,title_npiso,aggregate_by):
    
    if variable.startswith("area") and variable.endswith("cons"):
        # =====================================================================
        Title_FontSize = 6.1                                                    # Tamano de letra del nombre de la variable presentada en el mapa
        Cbar_Scale = 1.000                                                      # Espaciado de los ticks del cbar
        Cbar_LabelSize = 4                                                      # Tamano de letra de los ticks del cbar
        Cbar_FontSize = 6.5                                                     # Tamano de letra del titulo del cbar
        Cientific = 0                                                           # Notacion cientifica 1: si 0: no
        Cmap_text = '% Área construida'                                         # Titulo de cbar
        title1 = '% Área construida de la manzana'                              # Titulo variable del mapa
        title2 = 'para la taxonomía '+ Txn_Rep                                  # Titulo variable del mapa
        viridis_inverted = 'Reds'                                               # cbar color
        # =====================================================================
    elif variable.startswith("area") and variable.endswith("cons2"):
        # =====================================================================
        Title_FontSize = 6.5                                                    # Tamano de letra del nombre de la variable presentada en el mapa
        Cbar_Scale = 0.500                                                      # Espaciado de los ticks del cbar
        Cbar_LabelSize = 4                                                      # Tamano de letra de los ticks del cbar
        Cbar_FontSize = 6.5                                                     # Tamano de letra del titulo del cbar
        Cientific = 0                                                           # Notacion cientifica 1: si 0: no
        Cmap_text = '% Área construida'                                         # Titulo de cbar
        title1 = '% Área construida de la manzana'                              # Titulo variable del mapa
        title2 = 'para edificaciones de ' + title_npiso                         # Titulo variable del mapa
        viridis_inverted = cm.get_cmap('plasma').reversed()                     # cbar color
        # =====================================================================
    elif variable.startswith("aal") and variable.endswith("mllr"):
        # =====================================================================
        Title_FontSize = 6.5                                                    # Tamano de letra del nombre de la variable presentada en el mapa
        Cbar_Scale = 0.500                                                      # Espaciado de los ticks del cbar
        Cbar_LabelSize = 4                                                      # Tamano de letra de los ticks del cbar
        Cbar_FontSize = 6.5                                                     # Tamano de letra del titulo del cbar
        Cientific = 1                                                           # Notacion cientifica 1: si 0: no
        Cmap_text = 'Pérdida anual promedio'                                    # Titulo de cbar
        if aggregate_by == "manzana":
            title1 = 'Pérdida anual promedio por'                               # Titulo variable del mapa
            title2 = 'manzana censal al millar'                                 # Titulo variable del mapa
        else:
            title1 = 'Pérdida anual promedio por'                               # Titulo variable del mapa
            title2 = 'sección urbana al millar'                                 # Titulo variable del mapa
        viridis_inverted = 'Reds'                                               # cbar color
        # =====================================================================
    elif variable.startswith("aal") and variable.endswith("cop"):
        # =====================================================================
        Title_FontSize = 5.4                                                    # Tamano de letra del nombre de la variable presentada en el mapa
        Cbar_Scale = 0.500                                                      # Espaciado de los ticks del cbar
        Cbar_LabelSize = 4                                                      # Tamano de letra de los ticks del cbar
        Cbar_FontSize = 6.5                                                     # Tamano de letra del titulo del cbar
        Cientific = 0                                                           # Notacion cientifica 1: si 0: no
        Cmap_text = 'Pérdida anual promedio'                                    # Titulo de cbar
        if aggregate_by == "manzana":
            title1 = 'Pérdida anual promedio por manzana'                       # Titulo variable del mapa
            title2 = 'censal en miles de millones de pesos COP'                 # Titulo variable del mapa
        else:
            title1 = 'Pérdida anual promedio por sección'                       # Titulo variable del mapa
            title2 = 'urbana en miles de millones de pesos COP'                 # Titulo variable del mapa
        viridis_inverted = 'Reds'                                               # cbar color
        # =====================================================================
    elif variable.startswith("aad") and variable.endswith("fallecidos_hab"):
        # =====================================================================
        Title_FontSize = 6.5                                                    # Tamano de letra del nombre de la variable presentada en el mapa
        Cbar_Scale = 0.500                                                      # Espaciado de los ticks del cbar
        Cbar_LabelSize = 4                                                      # Tamano de letra de los ticks del cbar
        Cbar_FontSize = 6.5                                                     # Tamano de letra del titulo del cbar
        Cientific = 0                                                           # Notacion cientifica 1: si 0: no
        Cmap_text = 'Fallecidos anuales promedio'                               # Titulo de cbar
        if aggregate_by == "manzana":
            title1 = 'Fallecidos anuales promedio'                              # Titulo variable del mapa
            title2 = 'por manzana censal'                                       # Titulo variable del mapa
        else:
            title1 = 'Fallecidos anuales promedio'                              # Titulo variable del mapa
            title2 = 'por sección urbana'                                       # Titulo variable del mapa
        viridis_inverted = 'Reds'                                               # cbar color
        # =====================================================================
    elif variable.startswith("aad") and variable.endswith("fallecidos_100m_hab"):
        # =====================================================================
        Title_FontSize = 5.5                                                    # Tamano de letra del nombre de la variable presentada en el mapa
        Cbar_Scale = 1.000                                                      # Espaciado de los ticks del cbar
        Cbar_LabelSize = 4                                                      # Tamano de letra de los ticks del cbar
        Cbar_FontSize = 6.5                                                     # Tamano de letra del titulo del cbar
        Cientific = 0                                                           # Notacion cientifica 1: si 0: no
        Cmap_text = 'Fallecidos anuales promedio'                               # Titulo de cbar
        if aggregate_by == "manzana":
            title1 = 'Fallecidos anuales promedio por cada'                     # Titulo variable del mapa
            title2 = '100k habitantes por manzana censal'                       # Titulo variable del mapa
        else:
            title1 = 'Fallecidos anuales promedio por cada'                     # Titulo variable del mapa
            title2 = '100k habitantes por sección urbana'                       # Titulo variable del mapa
        viridis_inverted = 'Reds'                                               # cbar color
        # =====================================================================
    elif variable.startswith("aai") and variable.endswith("heridos_hab"):
        # =====================================================================
        Title_FontSize = 6.5                                                    # Tamano de letra del nombre de la variable presentada en el mapa
        Cbar_Scale = 0.500                                                      # Espaciado de los ticks del cbar
        Cbar_LabelSize = 4                                                      # Tamano de letra de los ticks del cbar
        Cbar_FontSize = 6.5                                                     # Tamano de letra del titulo del cbar
        Cientific = 0                                                           # Notacion cientifica 1: si 0: no
        Cmap_text = 'Heridos anuales promedio'                                  # Titulo de cbar
        if aggregate_by == "manzana":
            title1 = 'Herido anuales promedio'                                  # Titulo variable del mapa
            title2 = 'por manzana censal'                                       # Titulo variable del mapa
        else:
            title1 = 'Herido anuales promedio'                                  # Titulo variable del mapa
            title2 = 'por sección urbana'                                       # Titulo variable del mapa
        viridis_inverted = 'PuRd'                                               # cbar color
        # =====================================================================
    elif variable.startswith("aai") and variable.endswith("heridos_100m_hab"):
        # =====================================================================
        Title_FontSize = 5.7                                                    # Tamano de letra del nombre de la variable presentada en el mapa
        Cbar_Scale = 1.000                                                      # Espaciado de los ticks del cbar
        Cbar_LabelSize = 4                                                      # Tamano de letra de los ticks del cbar
        Cbar_FontSize = 6.5                                                     # Tamano de letra del titulo del cbar
        Cientific = 0                                                           # Notacion cientifica 1: si 0: no
        Cmap_text = 'Heridos anuales promedio'                                  # Titulo de cbar
        if aggregate_by == "manzana":
            title1 = 'Heridos anuales promedio por cada'                        # Titulo variable del mapa
            title2 = '100k habitantes por manzana censal'                       # Titulo variable del mapa
        else:
            title1 = 'Heridos anuales promedio por cada'                        # Titulo variable del mapa
            title2 = '100k habitantes por sección urbana'                       # Titulo variable del mapa
        viridis_inverted = 'PuRd'                                               # cbar color
        # =====================================================================
    elif variable.startswith("aac") and variable.endswith("colapso_hab"):
        # =====================================================================
        Title_FontSize = 5.4                                                    # Tamano de letra del nombre de la variable presentada en el mapa
        Cbar_Scale = 1.000                                                      # Espaciado de los ticks del cbar
        Cbar_LabelSize = 4                                                      # Tamano de letra de los ticks del cbar
        Cbar_FontSize = 6.5                                                     # Tamano de letra del titulo del cbar
        Cientific = 0                                                           # Notacion cientifica 1: si 0: no
        Cmap_text = 'Ocupantes en edificios colapsados'                         # Titulo de cbar
        if aggregate_by == "manzana":
            title1 = 'Ocupantes en edificios colapsados'                        # Titulo variable del mapa
            title2 = 'edificios colapsados por manzana censal'                  # Titulo variable del mapa
        else:
            title1 = 'Ocupantes en edificios colapsados'                        # Titulo variable del mapa
            title2 = 'edificios colapsados por sección urbana'                  # Titulo variable del mapa
        viridis_inverted = 'GnBu'                                               # cbar color
        # =====================================================================
    elif variable.startswith("aac") and variable.endswith("colapso_no_edis"):
        # =====================================================================
        Title_FontSize = 5.9                                                    # Tamano de letra del nombre de la variable presentada en el mapa
        Cbar_Scale = 0.500                                                      # Espaciado de los ticks del cbar
        Cbar_LabelSize = 4                                                      # Tamano de letra de los ticks del cbar
        Cbar_FontSize = 6.5                                                     # Tamano de letra del titulo del cbar
        Cientific = 0                                                           # Notacion cientifica 1: si 0: no
        Cmap_text = 'Edificios colapsados'                                      # Titulo de cbar
        if aggregate_by == "manzana":
            title1 = 'Número anual promedio de edificios'                       # Titulo variable del mapa
            title2 = 'colapsados por manzana censal'                            # Titulo variable del mapa
        else:
            title1 = 'Número anual promedio de edificios'                       # Titulo variable del mapa
            title2 = 'colapsados por sección urbana'                            # Titulo variable del mapa
        viridis_inverted = 'GnBu'                                               # cbar color
        # =====================================================================
    elif variable.startswith("dmg") and variable.endswith("4"):
        # =====================================================================
        Title_FontSize = 6.5                                                    # Tamano de letra del nombre de la variable presentada en el mapa
        Cbar_Scale = 0.030                                                      # Espaciado de los ticks del cbar
        Cbar_LabelSize = 4                                                      # Tamano de letra de los ticks del cbar
        Cbar_FontSize = 6.5                                                     # Tamano de letra del titulo del cbar
        Cientific = 0                                                           # Notacion cientifica 1: si 0: no
        Cmap_text = 'Edificaciones colapsadas [%]'                              # Titulo de cbar
        title1 = 'Edificaciones en estado de'                                   # Titulo variable del mapa
        title2 = 'colapso por manzana censal [%]'                               # Titulo variable del mapa
        viridis_inverted = cm.get_cmap('plasma').reversed()                     # cbar color
        # =====================================================================
    elif variable.startswith("dmg") and variable.endswith("3"):
        # =====================================================================
        Title_FontSize = 6.5                                                    # Tamano de letra del nombre de la variable presentada en el mapa
        Cbar_Scale = 0.030                                                      # Espaciado de los ticks del cbar
        Cbar_LabelSize = 4                                                      # Tamano de letra de los ticks del cbar
        Cbar_FontSize = 6.5                                                     # Tamano de letra del titulo del cbar
        Cientific = 0                                                           # Notacion cientifica 1: si 0: no
        Cmap_text = 'Edificaciones con daño severo [%]'                         # Titulo de cbar
        title1 = 'Edificaciones en daño'                                        # Titulo variable del mapa
        title2 = 'severo por manzana censal [%]'                                # Titulo variable del mapa
        viridis_inverted = cm.get_cmap('plasma').reversed()                     # cbar color
        # =====================================================================
    elif variable.startswith("pae") and variable.endswith("cop"):
        # =====================================================================
        Title_FontSize = 6.5                                                    # Tamano de letra del nombre de la variable presentada en el mapa
        Cbar_Scale = 0.003                                                      # Espaciado de los ticks del cbar
        Cbar_LabelSize = 4                                                      # Tamano de letra de los ticks del cbar
        Cbar_FontSize = 6.5                                                     # Tamano de letra del titulo del cbar
        Cientific = 0                                                           # Notacion cientifica 1: si 0: no
        Cmap_text = 'PAE [COP Millones]'                                        # Titulo de cbar
        title1 = 'Pérdida anual esperada en'                                    # Titulo variable del mapa
        title2 = 'millones de pesos COP'                                        # Titulo variable del mapa
        viridis_inverted = cm.get_cmap('viridis').reversed()                    # cbar color
        # =====================================================================
    elif variable.startswith("pae") and variable.endswith("prc"):
        # =====================================================================
        Title_FontSize = 6.5                                                    # Tamano de letra del nombre de la variable presentada en el mapa
        Cbar_Scale = 0.500                                                      # Espaciado de los ticks del cbar
        Cbar_LabelSize = 4                                                      # Tamano de letra de los ticks del cbar
        Cbar_FontSize = 6.5                                                     # Tamano de letra del titulo del cbar
        Cientific = 0                                                           # Notacion cientifica 1: si 0: no
        Cmap_text = 'PAE [‰]'                                                   # Titulo de cbar
        title1 = 'Pérdida anual esperada en'                                    # Titulo variable del mapa
        title2 = '‰ (PAE/Valor espuesto)'                                       # Titulo variable del mapa
        viridis_inverted = cm.get_cmap('viridis').reversed()                    # cbar color
        # =====================================================================
    
    # Generar dataframe data_figura38
    # Para eso necesito:
        # 1. Modelo de exposicion del municipio (informar en guia el nombre de las columnas)
        
    # Codigo de las manzanas corregido

    cod_mnzdef_model = []
    for mnz in Modelo_Expo.cod_mnz:
        if mnz[0].isalpha():
            mnz2 = str(mnz)
            cod_mnzdef_model.append(mnz2[1::])   
        else:
            mnz2 = str(mnz)
            cod_mnzdef_model.append(mnz2)  
            
    Modelo_Expo.cod_mnz = cod_mnzdef_model

    cod_sccdef_model = []
    for scc in Modelo_Expo.cod_secc:
        if scc[0].isalpha():
            cod_sccdef_model.append(scc[1::])
        else:
            cod_sccdef_model.append(scc)
               
    Modelo_Expo.cod_secc = cod_sccdef_model
    
    # Obtener taxonomia representativa
    Expo_groupby_txn = Modelo_Expo.groupby('tipologia')['area_cons'] 
    Expo_groupby_mnz = Modelo_Expo.groupby('cod_mnz')['area_cons'].sum()

    Area_constr_txn = Expo_groupby_txn.sum()
    df_area = pd.DataFrame({'tipologia':Area_constr_txn.index,'area':Area_constr_txn})
    Txn_Rep = df_area.loc[df_area.area == df_area.area.max()].values[0][0]          
    # Obtener el % de area construida de las manzanas con esa taxonomia representativa
    df_mnz_txn = Modelo_Expo.loc[Modelo_Expo.tipologia == Txn_Rep]
    df_mnz_groupbymnz = df_mnz_txn.groupby('cod_mnz')['area_cons'].sum()
    df_final = pd.DataFrame({'cod_mnz2':df_mnz_groupbymnz.index,'area':df_mnz_groupbymnz})
    df_final = df_final.merge(Expo_groupby_mnz, left_on='cod_mnz2', right_on='cod_mnz', how='left')
    df_final['area_cons'] = (df_final.area/df_final.area_cons)*100 

    # Geopandas para definir los limites del mapa
    map_data_limits = df_final.merge(manzana_shp, left_on='cod_mnz2', right_on='COD_DANE', how='left')
    map_data_limits_df = map_data_limits.to_csv("Map_Limits.csv",index=False)

    # Generar mapa de manzanas
    if aggregate_by == "manzana":
        figura = mapa_gen_mnz(Cientific,Separa_x,Separa_y,COD_mun,CP_Name,manzana_shp,seccion_shp,area_shp,map_data,map_data_limits_df,variable,viridis_inverted,Cmap_text,title1,title2,Title_FontSize,Cbar_Scale,Cbar_LabelSize,Cbar_FontSize,User_min_lon,User_max_lon,User_min_lat,User_max_lat)
    else:
        figura = mapa_gen_scc(Cientific,Separa_x,Separa_y,COD_mun,CP_Name,seccion_shp,area_shp,map_data,map_data_limits_df,variable,viridis_inverted,Cmap_text,title1,title2,Title_FontSize,Cbar_Scale,Cbar_LabelSize,Cbar_FontSize,User_min_lon,User_max_lon,User_min_lat,User_max_lat)

    
    return figura

#%% ====== Funcion - Fichas tecnicas ==========================================

def Taxo_Description(df_expotax):
    taxo_description = []
    for txn in df_expotax.taxonomy:
        parte = txn.split('/')
        if parte[0] == 'VG':
            if parte[1] == 'CE':
                if parte[2] == 'DU':
                    taxo_description.append('Cerchas de material vegetal (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Cerchas de material vegetal (no ingenieril)')
                else:
                    taxo_description.append('Cerchas de material vegetal (NI)')
        elif parte[0] == 'AC':
            if parte[1] == 'CE':
                if parte[2] == 'DU':
                    taxo_description.append('Cerchas de acero (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Cerchas de acero (no ingenieril)')
                else:
                    taxo_description.append('Cerchas de acero (NI)')
            if parte[1] == 'MD':
                if parte[2] == 'DU':
                    taxo_description.append('Muros delgados en acero (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Muros delgados en acero (no ingenieril)')
                else:
                    taxo_description.append('Muros delgados en acero(NI)')
            elif parte[1] == 'PRM':
                if parte[2] == 'DU':
                    taxo_description.append('Pórticos resistentes a momento de acero (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Pórticos resistentes a momento de acero (no ingenieril)')
                else:
                    taxo_description.append('Pórticos resistentes a momento de acero (NI)')
            elif parte[1] == 'PA':
                if parte[2] == 'DU':
                    taxo_description.append('Pórticos arriostrados de acero (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Pórticos arriostrados de acero (no ingenieril)')
                else:
                    taxo_description.append('Pórticos arriostrados de acero (NI)')
            elif parte[1] == 'PI':
                if parte[2] == 'DU':
                    taxo_description.append('Péndulo invertido de acero (ingenieril)')
                else:
                    taxo_description.append('Péndulo invertido de acero (no ingenieril)')
        elif parte[0] == 'AD':
            taxo_description.append('Adobe no-reforzado')
        elif parte[0] == 'BQ':
            taxo_description.append('Bahareque no-reforzado')
        elif parte[0] == 'MZ':
            if parte[1] == 'OT':
                taxo_description.append('Madera-zinc')
            else:
                taxo_description.append('Madera-zinc no-reforzada')
        elif parte[0] == 'MX':
            taxo_description.append('Mixto (No ingenieril)')
        elif parte[0] == 'CR':
            if parte[1] == 'MR':
                if parte[2] == 'DU':
                    taxo_description.append('Muros de concreto reforzado (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Muros de concreto reforzado (no ingenieril)')
                elif parte[2] == 'NI':
                    taxo_description.append('Muros de concreto reforzado (NI)')
            elif parte[1] == 'MD':
                if parte[2] == 'DU':
                    taxo_description.append('Muros delgados de concreto reforzado (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Muros delgados de concreto reforzado (no ingenieril)')
                elif parte[2] == 'NI':
                    taxo_description.append('Muros delgados de concreto reforzado (NI)')
            elif parte[1] == 'PRM':
                if parte[2] == 'DU':
                    taxo_description.append('Pórticos resistentes a momento de concreto reforzado (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Pórticos resistentes a momento de concreto reforzado (no ingenieril)')
                elif parte[2] == 'NI':
                    taxo_description.append('Pórticos resistentes a momento de concreto reforzado (NI)')
            elif parte[1] == 'PRMM':
                if parte[2] == 'DU':
                    taxo_description.append('Pórticos resistentes a momento de concreto reforzado con relleno en mampostería (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Pórticos resistentes a momento de concreto reforzado con relleno en mampostería (no ingenieril)')
                elif parte[2] == 'NI':
                    taxo_description.append('Pórticos resistentes a momento de concreto reforzado con relleno en mampostería (NI)')
            elif parte[1] == 'PA':
                if parte[2] == 'DU':
                    taxo_description.append('Pórticos arriostrados de concreto reforzado (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Pórticos arriostrados de concreto reforzado (no ingenieril)')
                elif parte[2] == 'NI':
                    taxo_description.append('Pórticos arriostrados de concreto reforzado (NI)')
            elif parte[1] == 'LC':
                taxo_description.append('Losa-Columna de concreto reforzado (ingenieril)')
            elif parte[1] == 'SC':
                if parte[2] == 'DU':
                    taxo_description.append('Sistema combinado (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Sistema combinado (no ingenieril)')
                elif parte[2] == 'NI':
                    taxo_description.append('Sistema combinado (NI)')
        elif parte[0] == 'MA':
            if parte[1] == 'MR':
                if parte[2] == 'DU':
                    taxo_description.append('Muros de mampostería reforzada (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Muros de mampostería reforzada (no ingenieril)')
                elif parte[2] == 'NI':
                    taxo_description.append('Muros de mampostería reforzada (NI)')
            elif parte[1] == 'PRM':
                if parte[2] == 'DU':
                    taxo_description.append('Pórticos resistentes a momento de mampostería reforzada (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Pórticos resistentes a momento de mampostería reforzada (no ingenieril)')
                elif parte[2] == 'NI':
                    taxo_description.append('Pórticos resistentes a momento de mampostería reforzada (NI)')
            elif parte[1] == 'MNR':
                taxo_description.append('Mampostería no-reforzada')
            elif parte[1] == 'MPC':
                taxo_description.append('Mampostería parcialmente confinada')
            elif parte[1] == 'MC':
                if parte[2] == 'DU':
                    taxo_description.append('Mampostería confinada (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Mampostería confinada (no ingenieril)')
                elif parte[2] == 'NI':
                    taxo_description.append('Mampostería confinada (NI)')
        elif parte[0] == 'MC':
            if parte[1] == 'MR':
                if parte[2] == 'DU':
                    taxo_description.append('Muros de mampostería en bloque de concreto reforzada (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Muros de mampostería en bloque de concreto reforzada (no ingenieril)')
                elif parte[2] == 'NI':
                    taxo_description.append('Muros de mampostería en bloque de concreto reforzada (NI)')
            elif parte[1] == 'PRM':
                if parte[2] == 'DU':
                    taxo_description.append('Pórticos resistentes a momento de mampostería en bloque de concreto reforzada (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Pórticos resistentes a momento de mampostería en bloque de concreto reforzada (no ingenieril)')
                elif parte[2] == 'NI':
                    taxo_description.append('Pórticos resistentes a momento de mampostería en bloque de concreto reforzada (NI)')
            elif parte[1] == 'MNR':
                taxo_description.append('Mampostería en bloque de concreto no-reforzada')
            elif parte[1] == 'MPC':
                taxo_description.append('Mampostería en bloque de concreto parcialmente confinada')
            elif parte[1] == 'MC':
                if parte[2] == 'DU':
                    taxo_description.append('Mampostería en bloque de concreto confinada (ingenieril)')
                elif parte[2] == 'ND':
                    taxo_description.append('Mampostería en bloque de concreto confinada (no ingenieril)')
                elif parte[2] == 'NI':
                    taxo_description.append('Mampostería en bloque de concreto confinada (NI)')
        elif parte[0]=='MD':
            taxo_description.append('Madera no-reforzada')
        elif parte[0] == 'PMC':
            taxo_description.append('Prefabricado (madera-concreto) no-reforzado')
        elif parte[0] == 'TA':
            taxo_description.append('Tapia pisada no-reforzada')
        elif parte[0] == 'MP':
            taxo_description.append('Mampostería de piedra no-reforzada')
        elif parte[0] == 'PMF':
            taxo_description.append('Prefabricado: metálico-fibrocemento')
        elif parte[0] == 'NI':
            taxo_description.append('Taxonomía no identificada')    
        
    return taxo_description

def mapa_gen_Infograph(Cientific,area_shp,manzana_shp,map_data,variable,viridis_inverted,Name_File,Cbar_Scale,Cbar_LabelSize,Cbar_FontSize,Cmap_text):
    
    fig, ax = plt.subplots(figsize=(6,5))
    plt.subplots_adjust(left=0.05, right=0.98, top=0.98, bottom=0.06)

    # .... Cargar los shape files .............................................
    valor_maximo = map_data[variable].max()
    
    if valor_maximo < 1:
        norm = Normalize(vmin=0, vmax=1.0)
        map_data.plot(column=variable, ax=ax, edgecolor='grey', alpha=1.0, cmap=viridis_inverted, linewidth=0.4, norm=norm)
    else:
        map_data.plot(column=variable, ax=ax, edgecolor='grey', alpha=1.0, cmap=viridis_inverted, linewidth=0.4)
        
    ax.set_axis_off()
    
    manzana_shp.plot(ax=ax, edgecolor='grey', facecolor="none", alpha=1.0, linewidth=0.4)
    area_shp.plot(ax=ax, edgecolor='black', facecolor="none", alpha=1.0, linewidth=0.6)
    
    if valor_maximo < 1:
        cax = fig.add_axes([0.75, 0.15, 0.02, 0.3]) # [left, bottom, width, height]
        cbar = fig.colorbar(cm.ScalarMappable(norm=norm, cmap=viridis_inverted), cax=cax, orientation='vertical')
        valores = np.linspace(0, 1.0, 3)
        
    else:
        norm = Normalize(vmin=np.floor(map_data[variable].min()), vmax=np.ceil(map_data[variable].max()))
    
        cax = fig.add_axes([0.75, 0.15, 0.02, 0.3]) # [left, bottom, width, height]
        cbar = fig.colorbar(cm.ScalarMappable(norm=norm, cmap=viridis_inverted), cax=cax, orientation='vertical')
    
        valores = np.linspace(np.floor(map_data[variable].min()), np.ceil(map_data[variable].max()), 6)
    

    if Cientific == 1:
        formatter = FuncFormatter(scientific_formatter)
        cbar.set_ticks(np.round(valores[:] / Cbar_Scale) * Cbar_Scale)
        cbar.ax.xaxis.set_major_formatter(formatter)
        cbar.ax.tick_params(axis='x', labelsize=Cbar_LabelSize)
    else:
        cbar.set_ticks(np.round(valores[:] / Cbar_Scale) * Cbar_Scale)
        cbar.ax.tick_params(axis='x', labelsize=Cbar_LabelSize)
    
    ax.text(1.05, 0.28, Cmap_text, transform=ax.transAxes,ha='center', va='center', fontsize=9,rotation=90)  
    
    # Ajustar márgenes para reducir el espacio en blanco
    plt.subplots_adjust(left=0.1, right=0.9, top=0.95, bottom=0.05)

    plt.savefig(os.path.join(os.path.join(os.getcwd(),"css"), Name_File), dpi=300, bbox_inches='tight', pad_inches=0)
    plt.show()

def Diagrama_no_edis(categorias,Colapsos_prc,Name_File):

    # Utilizar el mapa de colores 'plasma' invertido
    colors = plt.cm.get_cmap('bone').reversed()(np.linspace(0.2, 1, len(categorias)))
    
    # Crear el diagrama de barras con el nuevo mapa de colores y sin marco
    fig, ax = plt.subplots()
    bars = ax.bar(range(len(categorias)), Colapsos_prc, color=colors)
    
    # Añadir los valores encima de cada barra
    for bar in bars:
        yval = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2, yval, f'{yval:.1f}%', ha='center', va='bottom',fontname='Calibri', fontsize=9)
    
    # Quitar ticks, títulos de los ejes y el marco del gráfico
    ax.set_xticks([])
    ax.set_yticks([])
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    
    # Añadir leyendas
    ax.legend(bars, categorias,  bbox_to_anchor=(0.5, -0.20), ncol=3, fontsize=9, loc='lower center', frameon=False, handlelength=1, handletextpad=0.4)
    
    
    # Ajustar márgenes para reducir el espacio en blanco
    plt.subplots_adjust(left=0.1, right=0.9, top=0.95, bottom=0.05)
    
    plt.savefig(os.path.join(os.path.join(os.getcwd(),"css"), Name_File), dpi=300, bbox_inches='tight', pad_inches=0)
    plt.show()
#%% ====== Funcion Exportar ===================================================
"""
-------------------------------------------------------------------------------
Funciones exportar archivos
-------------------------------------------------------------------------------
"""
def ExportarGraficos_Perdidas_Calibrar(canvas1, canvas2, canvas3):
    # Pregunta al usuario dónde desea guardar los archivos 
    directorio_destino = filedialog.askdirectory(title="Selecciona un directorio para guardar los resultados")
    if directorio_destino:
        # Guardar el primer gráfico
        filename1 = os.path.join(directorio_destino, "Calibracion_CP.jpg")
        canvas1.figure.savefig(filename1, bbox_inches='tight', format='jpg', dpi=300)
        # Guardar el segundo gráfico
        if canvas2 is not None:
            filename2 = os.path.join(directorio_destino, "Calibracion_MNZ.jpg")
            canvas2.figure.savefig(filename2, bbox_inches='tight', format='jpg', dpi=300)
        # Guardar el tercer0 gráfico
        if canvas3 is not None:
            filename3 = os.path.join(directorio_destino, "Calibracion_TXN.jpg")
            canvas3.figure.savefig(filename3, bbox_inches='tight', format='jpg', dpi=300)
        # Informar al usuario que los gráficos se han guardado con éxito
        tk.messagebox.showinfo("Exportar resultados", "Los archivos se han guardado en:\n\n" + directorio_destino)
        
def ExportarGraficos_Perdidas_Dispersion(canvas1, canvas2, canvas3):
    # Pregunta al usuario dónde desea guardar los archivos 
    directorio_destino = filedialog.askdirectory(title="Selecciona un directorio para guardar los resultados")
    if directorio_destino:
        # Guardar el primer gráfico
        filename1 = os.path.join(directorio_destino, "Dispersion_CP.jpg")
        canvas1.figure.savefig(filename1, bbox_inches='tight', format='jpg', dpi=300)
        # Guardar el segundo gráfico
        if canvas2 is not None:
            filename2 = os.path.join(directorio_destino, "Dispersion_MNZ.jpg")
            canvas2.figure.savefig(filename2, bbox_inches='tight', format='jpg', dpi=300)
        # Guardar el tercer0 gráfico
        if canvas3 is not None:
            filename3 = os.path.join(directorio_destino, "Dispersion_TXN.jpg")
            canvas3.figure.savefig(filename3, bbox_inches='tight', format='jpg', dpi=300)
        # Informar al usuario que los gráficos se han guardado con éxito
        tk.messagebox.showinfo("Exportar resultados", "Los archivos se han guardado en:\n\n" + directorio_destino)
        
def ExportarResultados_Event_Based_Risk(Excedence_Curve,Table_Resu,Table_Resu_Txn,Figure_txn):
    
    # Pregunta al usuario dónde desea guardar los archivos 
    directorio_destino = filedialog.askdirectory(title="Selecciona un directorio para guardar los resultados")
    
    if directorio_destino:
        
        # Guardar curva de excedencia
        filename_ce = os.path.join(directorio_destino, "Curva_Excedencia.jpg")
        Excedence_Curve.figure.savefig(filename_ce, bbox_inches='tight', format='jpg', dpi=300)
        # Guardar tabla de resumen del municipio PAE
        Table_Resu(os.path.join(directorio_destino, "PAE_Municipio.xlsx"))
        # Guardar tabla de resumen por taxonomia PAE
        Table_Resu_Txn(os.path.join(directorio_destino, "PAE_Taxonomias.xlsx"))
        # Diagrama PAE por taxonomia
        Figure_txn.savefig(os.path.join(directorio_destino, "Diagrama_PAE_Txn.jpg"), dpi=300)

        # Informar al usuario que los gráficos se han guardado con éxito
        tk.messagebox.showinfo("Exportar resultados", "Los archivos se han guardado en (EXCEPTO MAPAS): \n\n" + directorio_destino)
        
def ExportarMapa_Event_Based_Risk(Fig_Mapa,File_Name):
    
    # Pregunta al usuario dónde desea guardar los archivos 
    directorio_destino = filedialog.askdirectory(title="Selecciona un directorio para guardar los resultados")
    
    if directorio_destino:
        
        # Guardar curva de excedencia
        filename = os.path.join(directorio_destino, File_Name)
        Fig_Mapa.figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
    
        # Informar al usuario que los gráficos se han guardado con éxito
        tk.messagebox.showinfo("Exportar Mapa", "El mapa se ha guardado en: \n\n" + directorio_destino)


def ExportarResultados_Event_Based_Damage(Table_DNO_CP,Table_DNO_Txn,Figura_DNO_Txn):
    
    # Pregunta al usuario dónde desea guardar los archivos 
    directorio_destino = filedialog.askdirectory(title="Selecciona un directorio para guardar los resultados")
    
    if directorio_destino:
        
        # Guardar diagrama de taxonomias
        Figura_DNO_Txn.savefig(os.path.join(directorio_destino, "Danos_Taxonomia.jpg"), dpi=300)
        # Guardar tabla de resumen del municipio 
        Table_DNO_CP(os.path.join(directorio_destino, "Edificaciones_DMG.xlsx"))
        # Guardar tabla de resumen por taxonomia 
        Table_DNO_Txn(os.path.join(directorio_destino, "Taxonomias_DMG.xlsx"))

        # Informar al usuario que los gráficos se han guardado con éxito
        tk.messagebox.showinfo("Exportar resultados", "Los archivos se han guardado en (EXCEPTO MAPAS): \n\n" + directorio_destino)

def ExportarMapa_Event_Based_Damage(Fig_Mapa,File_Name):
    
    # Pregunta al usuario dónde desea guardar los archivos 
    directorio_destino = filedialog.askdirectory(title="Selecciona un directorio para guardar los resultados")
    
    if directorio_destino:
        
        # Guardar curva de excedencia
        filename = os.path.join(directorio_destino, File_Name)
        Fig_Mapa.figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
    
        # Informar al usuario que los gráficos se han guardado con éxito
        tk.messagebox.showinfo("Exportar Mapa", "El mapa se ha guardado en: \n\n" + directorio_destino)


def ExportarResultados_MapasGeograficos(Fig_Area_Cons,Fig_Area_Cons_Pisos,Lab_Area_Cons_Pisos,Fig_PAP_mll,Fig_PAP_cop_mnz,Fig_ftl_100_mnz,Fig_fatalities_mnz,Fig_inj_100_mnz,Fig_injured_mnz,Fig_homeless_mnz,Fig_colapsed_mnz,Fig_PAP_mll_scc,Fig_PAP_cop_scc,Fig_ftl_100_scc,Fig_fatalities_scc,Fig_inj_100_scc,Fig_injured_scc,Fig_homeless_scc,Fig_colapsed_scc,Fig_dmg3,Fig_dmg4,Fig_paecop,Fig_paeprc):
    # Pregunta al usuario dónde desea guardar los archivos 
    directorio_destino = filedialog.askdirectory(title="Selecciona un directorio para guardar los resultados")
    
    if directorio_destino:
        
        # Guardar figura 1
        filename = os.path.join(directorio_destino, Fig_Area_Cons[1])
        Fig_Area_Cons[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
    
        # Guardar figura 2
        for index in range(len(Fig_Area_Cons_Pisos)):
            filename = os.path.join(directorio_destino, Lab_Area_Cons_Pisos[index])
            Fig_Area_Cons_Pisos[index].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 3
        filename = os.path.join(directorio_destino, Fig_PAP_mll[1])
        Fig_PAP_mll[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 4
        filename = os.path.join(directorio_destino, Fig_PAP_cop_mnz[1])
        Fig_PAP_cop_mnz[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 5
        filename = os.path.join(directorio_destino, Fig_ftl_100_mnz[1])
        Fig_ftl_100_mnz[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 6
        filename = os.path.join(directorio_destino, Fig_fatalities_mnz[1])
        Fig_fatalities_mnz[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 7
        filename = os.path.join(directorio_destino, Fig_inj_100_mnz[1])
        Fig_inj_100_mnz[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 8
        filename = os.path.join(directorio_destino, Fig_injured_mnz[1])
        Fig_injured_mnz[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 9
        filename = os.path.join(directorio_destino, Fig_homeless_mnz[1])
        Fig_homeless_mnz[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 10
        filename = os.path.join(directorio_destino, Fig_colapsed_mnz[1])
        Fig_colapsed_mnz[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 11
        filename = os.path.join(directorio_destino, Fig_PAP_mll_scc[1])
        Fig_PAP_mll_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 12
        filename = os.path.join(directorio_destino, Fig_PAP_cop_scc[1])
        Fig_PAP_cop_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 13
        filename = os.path.join(directorio_destino, Fig_ftl_100_scc[1])
        Fig_ftl_100_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 14
        filename = os.path.join(directorio_destino, Fig_fatalities_scc[1])
        Fig_fatalities_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 15
        filename = os.path.join(directorio_destino, Fig_inj_100_scc[1])
        Fig_inj_100_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 16
        filename = os.path.join(directorio_destino, Fig_injured_scc[1])
        Fig_injured_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 17
        filename = os.path.join(directorio_destino, Fig_homeless_scc[1])
        Fig_homeless_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 18
        filename = os.path.join(directorio_destino, Fig_colapsed_scc[1])
        Fig_colapsed_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 19
        filename = os.path.join(directorio_destino, Fig_dmg3[1])
        Fig_dmg3[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 20
        filename = os.path.join(directorio_destino, Fig_dmg4[1])
        Fig_dmg4[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 21
        filename = os.path.join(directorio_destino, Fig_paecop[1])
        Fig_paecop[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 22
        filename = os.path.join(directorio_destino, Fig_paeprc[1])
        Fig_paeprc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        
        # Informar al usuario que los gráficos se han guardado con éxito
        tk.messagebox.showinfo("Exportar Mapa", "Los mapas se han guardado con éxito en: \n\n" + directorio_destino)

def ExportarResultados_Generador(Fig_Area_Cons,Fig_Area_Cons_Pisos,Lab_Area_Cons_Pisos,Fig_PAP_mll,Fig_PAP_cop_mnz,Fig_ftl_100_mnz,Fig_fatalities_mnz,Fig_inj_100_mnz,Fig_injured_mnz,Fig_homeless_mnz,Fig_colapsed_mnz,Fig_PAP_mll_scc,Fig_PAP_cop_scc,Fig_ftl_100_scc,Fig_fatalities_scc,Fig_inj_100_scc,Fig_injured_scc,Fig_homeless_scc,Fig_colapsed_scc,Fig_dmg3,Fig_dmg4,Fig_paecop,Fig_paeprc,Table1_Danos,Table2_Danos,Figura1_Danos,Table1_Perdidas,Table2_Perdidas,Figura1_Perdidas,Figura2_Perdidas):
    # Pregunta al usuario dónde desea guardar los archivos 
    directorio_destino = filedialog.askdirectory(title="Selecciona un directorio para guardar los resultados")
    
    if directorio_destino:
        
        # Guardar figura 1
        filename = os.path.join(directorio_destino, Fig_Area_Cons[1])
        Fig_Area_Cons[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
    
        # Guardar figura 2
        for index in range(len(Fig_Area_Cons_Pisos)):
            filename = os.path.join(directorio_destino, Lab_Area_Cons_Pisos[index])
            Fig_Area_Cons_Pisos[index].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 3
        filename = os.path.join(directorio_destino, Fig_PAP_mll[1])
        Fig_PAP_mll[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 4
        filename = os.path.join(directorio_destino, Fig_PAP_cop_mnz[1])
        Fig_PAP_cop_mnz[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 5
        filename = os.path.join(directorio_destino, Fig_ftl_100_mnz[1])
        Fig_ftl_100_mnz[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 6
        filename = os.path.join(directorio_destino, Fig_fatalities_mnz[1])
        Fig_fatalities_mnz[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 7
        filename = os.path.join(directorio_destino, Fig_inj_100_mnz[1])
        Fig_inj_100_mnz[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 8
        filename = os.path.join(directorio_destino, Fig_injured_mnz[1])
        Fig_injured_mnz[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 9
        filename = os.path.join(directorio_destino, Fig_homeless_mnz[1])
        Fig_homeless_mnz[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 10
        filename = os.path.join(directorio_destino, Fig_colapsed_mnz[1])
        Fig_colapsed_mnz[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 11
        filename = os.path.join(directorio_destino, Fig_PAP_mll_scc[1])
        Fig_PAP_mll_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 12
        filename = os.path.join(directorio_destino, Fig_PAP_cop_scc[1])
        Fig_PAP_cop_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 13
        filename = os.path.join(directorio_destino, Fig_ftl_100_scc[1])
        Fig_ftl_100_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 14
        filename = os.path.join(directorio_destino, Fig_fatalities_scc[1])
        Fig_fatalities_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 15
        filename = os.path.join(directorio_destino, Fig_inj_100_scc[1])
        Fig_inj_100_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 16
        filename = os.path.join(directorio_destino, Fig_injured_scc[1])
        Fig_injured_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 17
        filename = os.path.join(directorio_destino, Fig_homeless_scc[1])
        Fig_homeless_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 18
        filename = os.path.join(directorio_destino, Fig_colapsed_scc[1])
        Fig_colapsed_scc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 19
        filename = os.path.join(directorio_destino, Fig_dmg3[1])
        Fig_dmg3[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 20
        filename = os.path.join(directorio_destino, Fig_dmg4[1])
        Fig_dmg4[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 21
        filename = os.path.join(directorio_destino, Fig_paecop[1])
        Fig_paecop[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 22
        filename = os.path.join(directorio_destino, Fig_paeprc[1])
        Fig_paeprc[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 23
        filename = os.path.join(directorio_destino, Figura1_Danos[1])
        Figura1_Danos[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 24
        filename = os.path.join(directorio_destino, Figura1_Perdidas[1])
        Figura1_Perdidas[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar figura 25
        filename = os.path.join(directorio_destino, Figura2_Perdidas[1])
        Figura2_Perdidas[0].figure.savefig(filename, bbox_inches='tight', format='jpg', dpi=300)
        
        # Guardar tabla 1
        Table1_Danos[0](os.path.join(directorio_destino, Table1_Danos[1]))
        
        # Guardar tabla 2
        Table2_Danos[0](os.path.join(directorio_destino, Table2_Danos[1]))
        
        # Guardar tabla 3
        Table1_Perdidas[0](os.path.join(directorio_destino, Table1_Perdidas[1]))
        
        # Guardar tabla 4
        Table2_Perdidas[0](os.path.join(directorio_destino, Table2_Perdidas[1]))
        
        
        # Informar al usuario que los gráficos se han guardado con éxito
        tk.messagebox.showinfo("Exportar Mapa", "Los resultados se han guardado con éxito en: \n\n" + directorio_destino)
        
def ExportarResultados_Ficha(pdf_base_path,CP_Name_FCH):
    
    # Pregunta al usuario dónde desea guardar los archivos 
    directorio_destino = filedialog.askdirectory(title="Selecciona un directorio para guardar los resultados")
    
    if directorio_destino:
        
        output_pdf_path = os.path.join(directorio_destino, "FichaTecnica_"+CP_Name_FCH+".pdf")
        
        packet = io.BytesIO()
        c = canvas.Canvas(packet)
        
        # Guardar el nuevo contenido en memoria
        c.save()
        packet.seek(0)

        # Leer el PDF base y el nuevo PDF
        pdf_base = PyPDF2.PdfReader(open(pdf_base_path, "rb"))
        new_pdf = PyPDF2.PdfReader(packet)

        # Crear un objeto de escritura para el PDF de salida
        pdf_writer = PyPDF2.PdfWriter()

        # Obtener la primera página del PDF base
        page = pdf_base.pages[0]
        
        # Verificar si el new_pdf tiene al menos una página
        if len(new_pdf.pages) > 0:
            # Obtener el contenido del nuevo PDF (solo una página)
            new_page = new_pdf.pages[0]
            
            # Superponer el nuevo contenido en la primera página del PDF base
            page.merge_page(new_page)

        # # Obtener el contenido del nuevo PDF (solo una página)
        # new_page = new_pdf.pages[0]

        # # Superponer el nuevo contenido en la primera página del PDF base
        # page.merge_page(new_page)

        # Agregar la página modificada al escritor de PDF
        pdf_writer.add_page(page)

        # Agregar las demás páginas del PDF base
        for page_num in range(1, len(pdf_base.pages)):
            page = pdf_base.pages[page_num]
            pdf_writer.add_page(page)

        # Guardar el PDF combinado
        with open(output_pdf_path, "wb") as output_pdf:
            pdf_writer.write(output_pdf)
        
        # Informar al usuario que el PDF se ha guardado con éxito
        tk.messagebox.showinfo("Exportar Resultados", "La ficha técnica del municipio se ha guardado en \n\n" + directorio_destino)

def Exportar_ManualUsuario():
    
    pdf_base_path = os.path.join(os.path.join(os.getcwd(),"css"), "Manual_Usuario.pdf")
    
    # Pregunta al usuario dónde desea guardar los archivos 
    directorio_destino = filedialog.askdirectory(title="Selecciona un directorio para guardar el manual del usuario")
    
    if directorio_destino:
        
        output_pdf_path = os.path.join(directorio_destino, "Manual_Usuario.pdf")
        
        packet = io.BytesIO()
        c = canvas.Canvas(packet)
        
        # Guardar el nuevo contenido en memoria
        c.save()
        packet.seek(0)

        # Leer el PDF base y el nuevo PDF
        pdf_base = PyPDF2.PdfReader(open(pdf_base_path, "rb"))
        new_pdf = PyPDF2.PdfReader(packet)

        # Crear un objeto de escritura para el PDF de salida
        pdf_writer = PyPDF2.PdfWriter()

        # Obtener la primera página del PDF base
        page = pdf_base.pages[0]
        
        # Verificar si el new_pdf tiene al menos una página
        if len(new_pdf.pages) > 0:
            # Obtener el contenido del nuevo PDF (solo una página)
            new_page = new_pdf.pages[0]
            
            # Superponer el nuevo contenido en la primera página del PDF base
            page.merge_page(new_page)

        # # Obtener el contenido del nuevo PDF (solo una página)
        # new_page = new_pdf.pages[0]

        # # Superponer el nuevo contenido en la primera página del PDF base
        # page.merge_page(new_page)

        # Agregar la página modificada al escritor de PDF
        pdf_writer.add_page(page)

        # Agregar las demás páginas del PDF base
        for page_num in range(1, len(pdf_base.pages)):
            page = pdf_base.pages[page_num]
            pdf_writer.add_page(page)

        # Guardar el PDF combinado
        with open(output_pdf_path, "wb") as output_pdf:
            pdf_writer.write(output_pdf)
        
        # Informar al usuario que el PDF se ha guardado con éxito
        tk.messagebox.showinfo("Exportar Resultados", "El manual se ha guardado con éxito en \n\n" + directorio_destino)

